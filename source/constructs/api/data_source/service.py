import asyncio
from datetime import datetime
from io import BytesIO
import json
import os
import random
import re
import time
import traceback
from time import sleep

import boto3
from fastapi import File, UploadFile
import openpyxl
from openpyxl.styles import Font, PatternFill
import pymysql
from botocore.exceptions import ClientError

from catalog.service import delete_catalog_by_account_region as delete_catalog_by_account
from catalog.service import delete_catalog_by_database_region as delete_catalog_by_database_region
from common.abilities import (convert_provider_id_2_database_type, convert_provider_id_2_name, query_all_vpc)
from common.constant import const
from common.enum import (MessageEnum,
                         ConnectionState,
                         Provider,
                         DataSourceType,
                         DatabaseType,
                         JDBCCreateType)
from common.exception_handler import BizException
from common.query_condition import QueryCondition
from db.models_data_source import (Account)
from discovery_job.service import can_delete_database as can_delete_job_database
from discovery_job.service import delete_account as delete_job_by_account
from discovery_job.service import delete_database as delete_job_database
from .jdbc_schema import list_jdbc_databases
from . import s3_detector, rds_detector, glue_database_detector, jdbc_detector, crud
from .schemas import (AccountInfo, AdminAccountInfo,
                      JDBCInstanceSource, JDBCInstanceSourceUpdate, JdbcSource,
                      ProviderResourceFullInfo, SourceNewAccount, SourceRegion,
                      SourceResourceBase,
                      SourceCoverage,
                      SourceGlueDatabaseBase,
                      JDBCInstanceSourceUpdateBase,
                      DataLocationInfo,
                      JDBCInstanceSourceBase,
                      JDBCInstanceSourceFullInfo)
from common.reference_parameter import logger, admin_account_id, admin_region, partition, admin_bucket_name

SLEEP_TIME = 5
SLEEP_MIN_TIME = 2
GRANT_PERMISSIONS_RETRIES = 10

# for delegated account(IT), admin account could use this role to list stackset in IT account.
# CloudFormation Role name is: ListOrganizationRole
_delegated_role_name = os.getenv("DelegatedRoleName", 'ListOrganizationRole')

# agent account roles for all operations by admin account
# CloudFormation Role names are: RoleForAdmin, APIRole, GlueDetectionJobRole
_agent_role_name = os.getenv('AgentRoleNameList', 'RoleForAdmin')


sts = boto3.client('sts')
""" :type : pyboto3.sts """

_jdbc_url_patterns = [
        r'jdbc:redshift://[\w.-]+:\d+/([\w-]+)',
        r'jdbc:redshift://[\w.-]+:\d+',
        r'jdbc:mysql://[\w.-]+:\d+/([\w-]+)',
        r'jdbc:mysql://[\w.-]+:\d+',
        r'jdbc:postgresql://[\w.-]+:\d+/([\w-]+)',
        r'jdbc:postgresql://[\w.-]+:\d+',
        r'jdbc:oracle:thin://@[\w.-]+:\d+/([\w-]+)',
        r'jdbc:oracle:thin://@[\w.-]+:\d+:\w+',
        r'jdbc:sqlserver://[\w.-]+:\d+;databaseName=([\w-]+)',
        r'jdbc:sqlserver://[\w.-]+:\d+;database=([\w-]+)']

__s3_client = boto3.client('s3')

def build_s3_targets(bucket, credentials, region, is_init):
    s3 = boto3.client('s3',
                      aws_access_key_id=credentials['AccessKeyId'],
                      aws_secret_access_key=credentials['SecretAccessKey'],
                      aws_session_token=credentials['SessionToken'],
                      region_name=region)
    """ :type : pyboto3.s3 """
    response = s3.list_objects(Bucket=bucket, Delimiter='/')
    if 'Contents' not in response and 'CommonPrefixes' not in response:
        logger.info(response)
        raise BizException(MessageEnum.SOURCE_S3_EMPTY_BUCKET.get_code(),
                           MessageEnum.SOURCE_S3_EMPTY_BUCKET.get_msg())
    logger.info(response)
    s3_targets = []
    if 'CommonPrefixes' in response:
        for common_prefix in response['CommonPrefixes']:
            s3_targets.append(
                {
                    "Path": f"s3://{bucket}/{common_prefix['Prefix']}",
                    "SampleSize": 20,
                    "Exclusions": __get_excludes_file_exts()
                }
            )
    if 'Contents' in response:
        if is_init:
            s3_targets.append(
                {
                    "Path": f"s3://{bucket}",
                    "SampleSize": 20,
                    "Exclusions": __get_excludes_file_exts()
                }
            )
        else:
            for content in response['Contents']:
                s3_targets.append(
                    {
                        "Path": f"s3://{bucket}/{content['Key']}",
                        "SampleSize": 20,
                        "Exclusions": __get_excludes_file_exts()
                    }
                )
    logger.info("build_s3_targets")
    logger.info(s3_targets)
    return s3_targets

def sync_s3_connection(account: str, region: str, bucket: str):
    glue_connection_name = f"{const.SOLUTION_NAME}-{DatabaseType.S3.value}-{bucket}"
    glue_database_name = f"{const.SOLUTION_NAME}-{DatabaseType.S3.value}-{bucket}"
    crawler_name = f"{const.SOLUTION_NAME}-{DatabaseType.S3.value}-{bucket}"
    if region != admin_region:
        raise BizException(MessageEnum.SOURCE_DO_NOT_SUPPORT_CROSS_REGION.get_code(),
                           MessageEnum.SOURCE_DO_NOT_SUPPORT_CROSS_REGION.get_msg())

    state = crud.get_s3_bucket_source_glue_state(account, region, bucket)
    logger.info("sync_s3_connection state is:")
    logger.info(state)
    if state == ConnectionState.PENDING.value:
        raise BizException(MessageEnum.SOURCE_CONNECTION_NOT_FINISHED.get_code(),
                           MessageEnum.SOURCE_CONNECTION_NOT_FINISHED.get_msg())
    elif state == ConnectionState.CRAWLING.value:
        raise BizException(MessageEnum.SOURCE_CONNECTION_CRAWLING.get_code(),
                           MessageEnum.SOURCE_CONNECTION_CRAWLING.get_msg())
    # else:
    #     delete_glue_connection(account, region, crawler_name, glue_database_name, glue_connection_name)
    try:
        # PENDING｜ACTIVE｜ERROR with message
        crud.set_s3_bucket_source_glue_state(account, region, bucket, ConnectionState.PENDING.value)
        iam_role_name = crud.get_iam_role(account)
        assumed_role = sts.assume_role(
            RoleArn=f"{iam_role_name}",
            RoleSessionName="glue-s3-connection"
        )
        credentials = assumed_role['Credentials']
        s3_targets = build_s3_targets(bucket, credentials, region, True)
        logger.info("sync_s3_connection rebuild s3 targets:")
        logger.info(s3_targets)
        glue = boto3.client('glue',
                            aws_access_key_id=credentials['AccessKeyId'],
                            aws_secret_access_key=credentials['SecretAccessKey'],
                            aws_session_token=credentials['SessionToken'],
                            region_name=region
                            )
        """ :type : pyboto3.glue """

        try:
            glue.get_database(Name=glue_database_name)
        except Exception as e:
            logger.info("sync_s3_connection glue get_database error, and create database")
            logger.info(e)
            response = glue.create_database(DatabaseInput={'Name': glue_database_name})
            logger.info(response)
        # wait for database creation, several seconds
        lakeformation = boto3.client('lakeformation',
                                     aws_access_key_id=credentials['AccessKeyId'],
                                     aws_secret_access_key=credentials['SecretAccessKey'],
                                     aws_session_token=credentials['SessionToken'],
                                     region_name=region)
        """ :type : pyboto3.lakeformation """
        crawler_role_arn = __gen_role_arn(account_id=account,
                                          region=region,
                                          role_name='GlueDetectionJobRole')

        # retry for grant permissions
        num_retries = GRANT_PERMISSIONS_RETRIES
        while num_retries > 0:
            try:
                response = lakeformation.grant_permissions(
                    Principal={
                        'DataLakePrincipalIdentifier': f"{crawler_role_arn}"
                    },
                    Resource={
                        'Database': {
                            'Name': glue_database_name
                        }
                    },
                    Permissions=['ALL'],
                    PermissionsWithGrantOption=['ALL']
                )
            except Exception as e:
                sleep(SLEEP_MIN_TIME)
                num_retries -= 1
            else:
                break
        else:
            raise BizException(MessageEnum.SOURCE_UNCONNECTED.get_code(), MessageEnum.SOURCE_UNCONNECTED.get_msg())
        try:
            gt_cr_response = glue.get_crawler(Name=crawler_name)
            logger.info(gt_cr_response)
            try:
                if not state or state == ConnectionState.ACTIVE.value or state == ConnectionState.UNSUPPORTED.value \
                        or state == ConnectionState.ERROR.value or state == ConnectionState.STOPPING.value:
                    up_cr_response = glue.update_crawler(
                        Name=crawler_name,
                        Role=crawler_role_arn,
                        DatabaseName=glue_database_name,
                        Targets={
                            "S3Targets": build_s3_targets(bucket, credentials, region, False)
                        },
                        SchemaChangePolicy={
                            'UpdateBehavior': 'UPDATE_IN_DATABASE',
                            'DeleteBehavior': 'DELETE_FROM_DATABASE'
                        }
                    )
                    logger.info("update crawler:")
                    logger.info(up_cr_response)
            except Exception as e:
                logger.info("update_crawler s3 error")
                logger.info(str(e))
            logger.info(response)
        except Exception as e:
            response = glue.create_crawler(
                Name=crawler_name,
                Role=crawler_role_arn,
                DatabaseName=glue_database_name,
                Targets={
                    "S3Targets": s3_targets
                },
                Tags={
                    const.TAG_KEY: const.TAG_VALUE,
                    const.TAG_ADMIN_ACCOUNT_ID: admin_account_id
                },
            )
            logger.info(response)

        crud.create_s3_connection(account, region, bucket, glue_connection_name, glue_database_name, crawler_name)
    except Exception as err:
        crud.set_s3_bucket_source_glue_state(account, region, bucket, str(err))
        glue = boto3.client('glue',
                            aws_access_key_id=credentials['AccessKeyId'],
                            aws_secret_access_key=credentials['SecretAccessKey'],
                            aws_session_token=credentials['SessionToken'],
                            region_name=region
                            )
        """ :type : pyboto3.glue """

        try:
            glue.delete_crawler(crawler_name)
        except Exception as e:
            pass
        try:
            glue.delete_database(Name=glue_database_name)
        except Exception as e:
            pass
        try:
            glue.delete_connection(ConnectionName=glue_connection_name)
        except Exception as e:
            pass

        logger.error(traceback.format_exc())
        raise BizException(MessageEnum.SOURCE_CONNECTION_FAILED.get_code(),
                           str(err))


def sync_s3_connection_by_region(account: str, region: str):
    buckets = crud.list_s3_bucket_source_by_account(
        account_id=account,
        region=region,
        state=None
    )
    for bucket in buckets:
        try:
            sync_s3_connection(
                account=account,
                region=region,
                bucket=bucket
            )
        except Exception as err:
            logger.error(traceback.format_exc())


def check_subnet_has_outbound_route(credentials, region_name: str, subnet_id: str):
    ec2_client = boto3.client('ec2',
                              aws_access_key_id=credentials['AccessKeyId'],
                              aws_secret_access_key=credentials['SecretAccessKey'],
                              aws_session_token=credentials['SessionToken'],
                              region_name=region_name)
    response = ec2_client.describe_route_tables(
        Filters=[
            {
                'Name': 'association.subnet-id',
                'Values': [subnet_id]
            }
        ]
    )

    for route_table in response['RouteTables']:
        for route in route_table['Routes']:
            if route.get('DestinationCidrBlock') == '0.0.0.0/0':
                logger.info("Has 0.0.0.0/0 route.")
                return True
    return False


def check_link(credentials, region_name: str, vpc_id: str, rds_secret_id: str):
    ec2_client = boto3.client('ec2',
                              aws_access_key_id=credentials['AccessKeyId'],
                              aws_secret_access_key=credentials['SecretAccessKey'],
                              aws_session_token=credentials['SessionToken'],
                              region_name=region_name)
    endpoints = ec2_client.describe_vpc_endpoints(Filters=[{'Name': 'vpc-id', 'Values': [vpc_id]}])['VpcEndpoints']
    s3_endpoint_exists = False
    glue_endpoint_exists = False
    secret_endpoint_exists = False
    for endpoint in endpoints:
        if endpoint['ServiceName'] == f'com.amazonaws.{region_name}.s3':
            s3_endpoint_exists = True
        elif endpoint['ServiceName'] == f'com.amazonaws.{region_name}.glue':
            glue_endpoint_exists = True
        elif endpoint['ServiceName'] == f'com.amazonaws.{region_name}.secretsmanager':
            secret_endpoint_exists = True

    if not s3_endpoint_exists:
        raise BizException(MessageEnum.SOURCE_RDS_NO_VPC_S3_ENDPOINT.get_code(),
                           MessageEnum.SOURCE_RDS_NO_VPC_S3_ENDPOINT.get_msg())
    if not glue_endpoint_exists:
        raise BizException(MessageEnum.SOURCE_RDS_NO_VPC_GLUE_ENDPOINT.get_code(),
                           MessageEnum.SOURCE_RDS_NO_VPC_GLUE_ENDPOINT.get_msg())
    if rds_secret_id is not None and not secret_endpoint_exists:
        raise BizException(MessageEnum.SOURCE_RDS_NO_VPC_SECRET_MANAGER_ENDPOINT.get_code(),
                           MessageEnum.SOURCE_RDS_NO_VPC_SECRET_MANAGER_ENDPOINT.get_msg())
    logger.info("sync_rds_connection check_link :")
    logger.info(s3_endpoint_exists)
    logger.info(glue_endpoint_exists)
    logger.info(secret_endpoint_exists)

def sync_glue_database(account_id, region, glue_database_name):

    state = crud.get_glue_database_source_glue_state(account_id, region, glue_database_name)
    if state == ConnectionState.PENDING.value:
        raise BizException(MessageEnum.SOURCE_CONNECTION_NOT_FINISHED.get_code(),
                           MessageEnum.SOURCE_CONNECTION_NOT_FINISHED.get_msg())
    try:
        sqs = boto3.client(
            'sqs',
            region_name=region
        )
        message = {
            "detail": {
                "databaseName": glue_database_name,
                "databaseType": "glue",
                "accountId": account_id,
                "state": "Succeeded"
            },
            "region": region
        }
        sqs.send_message(
            QueueUrl=sqs.get_queue_url(QueueName=f"{const.SOLUTION_NAME}-Crawler")['QueueUrl'],
            MessageBody=json.dumps(message)
        )
        crud.set_glue_database_glue_state(account_id, region, glue_database_name, ConnectionState.PENDING.value)
    except Exception as err:
        crud.set_glue_database_glue_state(account_id, region, glue_database_name, str(err))


def sync_jdbc_connection(jdbc: JDBCInstanceSourceBase):
    account_id, region = __get_admin_info(jdbc)
    ec2_client, credentials = __ec2(account=account_id, region=region)
    glue_client = __glue(account=account_id, region=region)
    lakeformation_client = __lakeformation(account=account_id, region=region)
    crawler_role_arn = __gen_role_arn(account_id=account_id,
                                      region=region,
                                      role_name='GlueDetectionJobRole')
    # get connection name from sdp db
    source: JDBCInstanceSourceFullInfo = crud.get_jdbc_instance_source_glue(provider_id=jdbc.account_provider_id,
                                                                            account=jdbc.account_id,
                                                                            region=jdbc.region,
                                                                            instance_id=jdbc.instance_id)
    if not source or not source.glue_connection:
        raise BizException(MessageEnum.SOURCE_JDBC_CONNECTION_NOT_EXIST.get_code(),
                           MessageEnum.SOURCE_JDBC_CONNECTION_NOT_EXIST.get_msg())
    # query connection info from glue
    conn_response = glue_client.get_connection(Name=source.glue_connection)['Connection']
    # return glue_client.get_connection(Name=source.glue_connection)

    logger.debug(f"conn_response type is:{type(conn_response)}")
    logger.debug(f"conn_response is:{conn_response}")
    if conn_response.get('ConnectionProperties'):
        username = conn_response.get('ConnectionProperties', {}).get('USERNAME')
        password = conn_response.get('ConnectionProperties', {}).get('PASSWORD')
        secret = conn_response.get('ConnectionProperties', {}).get("SECRET_ID"),
        url = conn_response.get('ConnectionProperties', {}).get('JDBC_CONNECTION_URL'),
        jdbc_instance = JDBCInstanceSource(instance_id=jdbc.instance_id,
                                           account_provider_id=jdbc.account_provider_id,
                                           account_id=jdbc.account_id,
                                           region=jdbc.region,
                                           jdbc_connection_url=url[0],
                                           master_username=username,
                                           password=password,
                                           secret=secret[0])
        # jdbc_instance.jdbc_connection_url = url
    # condition_check(ec2_client, credentials, source.glue_state, conn_response['PhysicalConnectionRequirements'])
        sync(glue_client,
             lakeformation_client,
             credentials,
             crawler_role_arn,
             jdbc_instance,
             source.jdbc_connection_schema)
    else:
        raise BizException(MessageEnum.BIZ_UNKNOWN_ERR.get_code(),
                           MessageEnum.BIZ_UNKNOWN_ERR.get_msg())


def condition_check(ec2_client, credentials, state, connection: dict):
    security_groups = []
    # if not jdbc.master_username:
    #     raise BizException(MessageEnum.SOURCE_JDBC_NO_CREDENTIAL.get_code(),
    #                        MessageEnum.SOURCE_JDBC_NO_CREDENTIAL.get_msg())

    # if not jdbc.password and not jdbc.secret:
    #     raise BizException(MessageEnum.SOURCE_JDBC_NO_AUTH.get_code(),
    #                        MessageEnum.SOURCE_JDBC_NO_AUTH.get_msg())

    # if jdbc.password and jdbc.secret:
    #     raise BizException(MessageEnum.SOURCE_JDBC_DUPLICATE_AUTH.get_code(),
    #                        MessageEnum.SOURCE_JDBC_DUPLICATE_AUTH.get_msg())
    # res = crud.get_jdbc_instance_source_glue(jdbc.account_provider, jdbc.account_id, jdbc.region, jdbc.instance)
    if state == ConnectionState.PENDING.value:
        raise BizException(MessageEnum.SOURCE_CONNECTION_NOT_FINISHED.get_code(),
                           MessageEnum.SOURCE_CONNECTION_NOT_FINISHED.get_msg())

    elif state == ConnectionState.CRAWLING.value:
        raise BizException(MessageEnum.SOURCE_CONNECTION_CRAWLING.get_code(),
                           MessageEnum.SOURCE_CONNECTION_CRAWLING.get_msg())
    # credentials = None

    # try:
    #     iam_role_name = crud.get_iam_role(admin_account_id)

    #     assumed_role = sts.assume_role(
    #         RoleArn=f"{iam_role_name}",
    #         RoleSessionName="glue-jdbc-connection"
    #     )
    #     credentials = assumed_role['Credentials']
    # except Exception as err:
    #     raise BizException(MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_code(),
    #                        MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_msg())
    if credentials is None:
        raise BizException(MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_code(),
                           MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_msg())
    vpc_list = [vpc['VpcId'] for vpc in query_all_vpc(ec2_client)]
    # vpc_list = [vpc['VpcId'] for vpc in ec2_client.describe_vpcs()['Vpcs']]
    # vpcs = []
    # response = ec2_client.describe_vpcs()
    # vpcs.append(response['Vpcs'])
    # while 'NextToken' in response:
    #     response = ec2_client.describe_vpcs(NextToken=response['NextToken'])
    #     vpcs.append(response['Vpcs'])
    try:
        security_groups = ec2_client.describe_security_groups(Filters=[
            {'Name': 'vpc-id', 'Values': vpc_list},
            {'Name': 'group-name', 'Values': [const.SECURITY_GROUP_JDBC]}
        ])["SecurityGroups"]
    except Exception as e:
        logger.error(traceback.format_exc())
        raise BizException(MessageEnum.SOURCE_SECURITY_GROUP_NOT_EXISTS.get_code(),
                           MessageEnum.SOURCE_SECURITY_GROUP_NOT_EXISTS.get_msg())
    if not security_groups:
        raise BizException(MessageEnum.SOURCE_SECURITY_GROUP_NOT_EXISTS.get_code(),
                           MessageEnum.SOURCE_SECURITY_GROUP_NOT_EXISTS.get_msg())
    security_group = list(filter(lambda sg: sg["GroupName"] == const.SECURITY_GROUP_JDBC, security_groups))[0]
    inbound_route = security_group["IpPermissions"][0]
    if inbound_route["IpProtocol"] != "tcp" or inbound_route["FromPort"] != 0 or inbound_route["ToPort"] != 65535:
        raise BizException(MessageEnum.SOURCE_SG_INBOUND_ROUTE_NOT_VALID.get_code(),
                           MessageEnum.SOURCE_SG_INBOUND_ROUTE_NOT_VALID.get_msg())
    outbound_route = security_group["IpPermissionsEgress"][0]
    if outbound_route["IpProtocol"] != "-1" or not outbound_route["IpRanges"] or outbound_route["IpRanges"][0]["CidrIp"] != "0.0.0.0/0":
        raise BizException(MessageEnum.SOURCE_SG_OUTBOUND_ROUTE_NOT_VALID.get_code(),
                           MessageEnum.SOURCE_SG_OUTBOUND_ROUTE_NOT_VALID.get_msg())
    subnet = ec2_client.describe_subnets(SubnetIds=[connection["SubnetId"]])["Subnets"]
    if not subnet:
        raise BizException(MessageEnum.SOURCE_SUBNET_JDBC_NOT_EXISTS.get_code(),
                           MessageEnum.SOURCE_SUBNET_JDBC_NOT_EXISTS.get_msg())
    if subnet[0]["MapPublicIpOnLaunch"]:
        raise BizException(MessageEnum.SOURCE_SUBNET_NOT_PRIVATE.get_code(),
                           MessageEnum.SOURCE_SUBNET_NOT_PRIVATE.get_msg())
    # if not ec2_client.describe_nat_gateways(Filters=[{'Name': 'subnet-id', 'Values': [jdbc.network_subnet_id]}])['NatGateways']:
    #     raise BizException(MessageEnum.SOURCE_SUBNET_NOT_CONTAIN_NAT.get_code(),
    #                        MessageEnum.SOURCE_SUBNET_NOT_CONTAIN_NAT.get_msg())
    #  contain gateway && subnet‘s route connected to ngw
    ngw_list = [item for item in ec2_client.describe_nat_gateways()['NatGateways'] if item['VpcId'] == subnet[0]['VpcId']]
    if not ngw_list:
        raise BizException(MessageEnum.SOURCE_VPC_NOT_CONTAIN_NAT.get_code(),
                           MessageEnum.SOURCE_VPC_NOT_CONTAIN_NAT.get_msg())
    if not ec2_client.describe_availability_zones(Filters=[{'Name': 'zone-name', 'Values': [connection["AvailabilityZone"]]}])["AvailabilityZones"]:
        raise BizException(MessageEnum.SOURCE_AVAILABILITY_ZONE_NOT_EXISTS.get_code(),
                           MessageEnum.SOURCE_AVAILABILITY_ZONE_NOT_EXISTS.get_msg())


def sync(glue, lakeformation, credentials, crawler_role_arn, jdbc: JDBCInstanceSource, schemas: str):
    jdbc_targets = []
    _, glue_database_name, crawler_name = __gen_resources_name(jdbc)
    state, glue_connection_name = crud.get_jdbc_connection_glue_info(jdbc.account_provider_id, jdbc.account_id, jdbc.region, jdbc.instance_id)
    if state == ConnectionState.CRAWLING.value:
        raise BizException(MessageEnum.SOURCE_CONNECTION_CRAWLING.get_code(),
                           MessageEnum.SOURCE_CONNECTION_CRAWLING.get_msg())
    jdbc_source = JdbcSource(connection_url=jdbc.jdbc_connection_url, username=jdbc.master_username, password=jdbc.password, secret_id=jdbc.secret)
    db_names = get_db_names_4_jdbc(jdbc_source, schemas)
    try:
        for db_name in db_names:
            trimmed_db_name = db_name.strip()
            if trimmed_db_name:
                jdbc_targets.append({
                    'ConnectionName': glue_connection_name,
                    'Path': f"{trimmed_db_name}/%"
                })

        if db_names:
            """ :type : pyboto3.glue """
            try:
                conn = glue.get_connection(Name=glue_connection_name)
                logger.debug(conn)
            except Exception as e:
                logger.error(traceback.format_exc())
                raise BizException(MessageEnum.SOURCE_CONNECTION_NOT_EXIST.get_code(), str(e))
            try:
                glue.get_database(Name=glue_database_name)
            except Exception as e:
                response = glue.create_database(DatabaseInput={'Name': glue_database_name})
                logger.debug(response)
            # """ :type : pyboto3.lakeformation """
            grant_lakeformation_permission(lakeformation, crawler_role_arn, glue_database_name, GRANT_PERMISSIONS_RETRIES)
            crawler_state = ConnectionState.CRAWLING.value
            try:
                response = glue.get_crawler(Name=crawler_name)
                try:
                    if not state or state == ConnectionState.ACTIVE.value or state == ConnectionState.UNSUPPORTED.value \
                            or state == ConnectionState.ERROR.value or state == ConnectionState.STOPPING.value:
                        up_cr_response = glue.update_crawler(
                            Name=crawler_name,
                            Role=crawler_role_arn,
                            DatabaseName=glue_database_name,
                            Targets={
                                'JdbcTargets': jdbc_targets,
                            },
                            SchemaChangePolicy={
                                'UpdateBehavior': 'UPDATE_IN_DATABASE',
                                'DeleteBehavior': 'DELETE_FROM_DATABASE'
                            }
                        )
                except Exception:
                    logger.error(traceback.format_exc())
                try:
                    st_cr_response = glue.start_crawler(
                        Name=crawler_name
                    )
                except Exception as e:
                    logger.error(traceback.format_exc())
                    crawler_state = str(e)
            except Exception:
                try:
                    response = glue.create_crawler(
                        Name=crawler_name,
                        Role=crawler_role_arn,
                        DatabaseName=glue_database_name,
                        Targets={
                            'JdbcTargets': jdbc_targets,
                        },
                        Tags={
                            const.TAG_KEY: const.TAG_VALUE,
                            const.TAG_ADMIN_ACCOUNT_ID: admin_account_id
                        },
                    )
                except Exception:
                    logger.error(traceback.format_exc())
                try:
                    start_response = glue.start_crawler(
                        Name=crawler_name
                    )
                    logger.debug(start_response)
                except Exception as e:
                    logger.error(traceback.format_exc())
                    crawler_state = str(e)

            crud.update_jdbc_connection(jdbc.account_provider_id,
                                        jdbc.account_id,
                                        jdbc.region,
                                        jdbc.instance_id,
                                        glue_database_name,
                                        crawler_name)
            crud.set_jdbc_connection_glue_state(
                jdbc.account_provider_id,
                jdbc.account_id,
                jdbc.region,
                jdbc.instance_id,
                crawler_state)
        else:
            crud.set_jdbc_connection_glue_state(jdbc.account_provider_id, jdbc.account_id, jdbc.region, jdbc.instance_id,
                                                MessageEnum.SOURCE_JDBC_JDBC_NO_DATABASE.get_msg())
            raise BizException(MessageEnum.SOURCE_JDBC_JDBC_NO_DATABASE.get_code(), MessageEnum.SOURCE_JDBC_JDBC_NO_DATABASE.get_msg())
    except Exception as e:
        logger.error(traceback.format_exc())
        crud.set_jdbc_connection_glue_state(jdbc.account_provider_id, jdbc.account_id, jdbc.region, jdbc.instance_id, str(e))
        glue = boto3.client('glue',
                            aws_access_key_id=credentials['AccessKeyId'],
                            aws_secret_access_key=credentials['SecretAccessKey'],
                            aws_session_token=credentials['SessionToken'],
                            region_name=jdbc.region
                            )
        """ :type : pyboto3.glue """
        try:
            glue.delete_crawler(crawler_name)
        except Exception:
            pass
        try:
            glue.delete_database(Name=glue_database_name)
        except Exception:
            pass
        try:
            glue.delete_connection(ConnectionName=glue_connection_name)
        except Exception:
            pass

        raise BizException(MessageEnum.SOURCE_CONNECTION_FAILED.get_code(),
                           str(e))

def grant_lakeformation_permission(lakeformation, crawler_role_arn, glue_database_name, num_retries):
    while num_retries > 0:
        try:
            response = lakeformation.grant_permissions(
                        Principal={
                            'DataLakePrincipalIdentifier': f"{crawler_role_arn}"
                        },
                        Resource={
                            'Database': {
                                'Name': glue_database_name
                            }
                        },
                        Permissions=['ALL'],
                        PermissionsWithGrantOption=['ALL']
                    )
        except Exception as e:
            logger.error(traceback.format_exc())
            sleep(SLEEP_MIN_TIME)
            num_retries -= 1
        else:
            break
    else:
        raise BizException(MessageEnum.SOURCE_UNCONNECTED.get_code(), MessageEnum.SOURCE_UNCONNECTED.get_msg())

def before_delete_glue_database(provider, account, region, name):
    glue_database = crud.get_glue_database_source(provider, account, region, name)
    if glue_database is None:
        raise BizException(MessageEnum.SOURCE_GLUE_DATABASE_NO_INSTANCE.get_code(),
                           MessageEnum.SOURCE_GLUE_DATABASE_NO_INSTANCE.get_msg())

    if not can_delete_job_database(account_id=account, region=region, database_type=DatabaseType.GLUE.value,
                                   database_name=name):
        raise BizException(MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_code(),
                           MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_msg())


# Delete third-party connection
def before_delete_jdbc_connection(provider_id, account, region, instance_id, database_type):
    jdbc_instance: JDBCInstanceSource = crud.get_jdbc_instance_source(provider_id, account, region, instance_id)
    if jdbc_instance is None:
        raise BizException(MessageEnum.SOURCE_JDBC_CONNECTION_NOT_EXIST.get_code(),
                           MessageEnum.SOURCE_JDBC_CONNECTION_NOT_EXIST.get_msg())
    # if rds_instance.glue_crawler is None:
    #     raise BizException(MessageEnum.SOURCE_RDS_NO_CRAWLER.get_code(),
    #                        MessageEnum.SOURCE_RDS_NO_CRAWLER.get_msg())
    # if rds_instance.glue_database is None:
    #     raise BizException(MessageEnum.SOURCE_RDS_NO_DATABASE.get_code(),
    #                        MessageEnum.SOURCE_RDS_NO_DATABASE.get_msg())
    # crawler, if crawling try to stop and raise, if pending raise directly
    # res = crud.get_jdbc_instance_source_glue(provider_id, account, region, instance_id)
    if jdbc_instance.glue_state == ConnectionState.PENDING.value:
        raise BizException(MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_code(),
                           MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_msg())
    elif jdbc_instance.glue_state == ConnectionState.CRAWLING.value:
        try:
            assume_account, assume_region = gen_assume_account(provider_id, account, region)
            # Stop the crawler
            __glue(account=assume_account, region=assume_region).stop_crawler(Name=jdbc_instance.glue_crawler)
        except Exception as e:
            logger.error(traceback.format_exc())
        raise BizException(MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_code(),
                           MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_msg())
    elif jdbc_instance.glue_state == ConnectionState.ACTIVE.value:
        # if job running, do not stop but raising
        if not can_delete_job_database(account_id=account, region=region, database_type=database_type,
                                       database_name=jdbc_instance.instance_id):
            raise BizException(MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_code(),
                               MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_msg())
    else:
        logger.info(f"delete jdbc connection: {account},{region},{database_type},{jdbc_instance.instance_id}")
        return jdbc_instance.glue_crawler

def gen_assume_account(provider_id, account, region):
    account = account if provider_id == Provider.AWS_CLOUD.value else admin_account_id
    region = region if provider_id == Provider.AWS_CLOUD.value else admin_region
    return account, region

def delete_glue_database(provider_id: int, account: str, region: str, name: str):
    before_delete_glue_database(provider_id, account, region, name)
    err = []
    # 1/3 delete job database
    try:
        delete_job_database(account_id=account, region=region, database_type=DatabaseType.GLUE.value, database_name=name)
    except Exception as e:
        err.append(str(e))
    # 2/3 delete catalog
    try:
        delete_catalog_by_database_region(database=name, region=region, type=DatabaseType.GLUE.value)
    except Exception as e:
        err.append(str(e))
    # 3/3 delete source
    s3_bucket = crud.get_glue_database_source(provider_id, account, region, name)
    glue = __glue(account=account, region=region)
    try:
        glue.delete_crawler(Name=s3_bucket.glue_crawler)
    except Exception as e:
        err.append(str(e))
    try:
        glue.delete_database(Name=s3_bucket.glue_database)
    except Exception as e:
        err.append(str(e))

    crud.delete_glue_database(provider_id, account, region, name)
    try:
        crud.update_glue_database_count(account, region)
    except Exception as e:
        err.append(str(e))

    if err:
        logger.error(traceback.format_exc())
        # raise BizException(MessageEnum.SOURCE_S3_CONNECTION_DELETE_ERROR.get_code(), err)

    return True

def delete_jdbc_connection(provider_id: int, account: str, region: str, instance_id: str, delete_catalog_only=False):
    database_type = convert_provider_id_2_database_type(provider_id)
    before_delete_jdbc_connection(provider_id, account, region, instance_id, database_type)
    assume_account, assume_region = gen_assume_account(provider_id, account, region)
    err = []
    # 1/3 delete job database
    try:
        if not delete_catalog_only:
            logger.info('delete_job_database start')
            delete_job_database(account_id=account, region=region, database_type=database_type, database_name=instance_id)
            logger.info('delete_job_database end')
    except Exception as e:
        logger.error(traceback.format_exc())
        err.append(str(e))
    # 2/3 delete catalog
    try:
        logger.info('delete_catalog_by_database_region start')
        delete_catalog_by_database_region(database=instance_id, region=region, type=database_type)
        logger.info('delete_catalog_by_database_region end')
    except Exception as e:
        logger.error(traceback.format_exc())
        err.append(str(e))

    if not delete_catalog_only:
        # 3/3 delete source
        jdbc_conn = crud.get_jdbc_instance_source(provider_id, account, region, instance_id)
        glue = __glue(account=assume_account, region=assume_region)
        if jdbc_conn.glue_crawler:
            try:
                logger.info(f'delete_crawler start:{assume_account, jdbc_conn.glue_crawler}')
                glue.delete_crawler(Name=jdbc_conn.glue_crawler)
                logger.info(f'delete_crawler end:{jdbc_conn.glue_crawler}')
            except Exception as e:
                logger.error(traceback.format_exc())
                err.append(str(e))
        if jdbc_conn.glue_database:
            try:
                glue.delete_database(Name=jdbc_conn.glue_database)
            except Exception as e:
                logger.error(traceback.format_exc())
                err.append(str(e))
        if jdbc_conn.glue_connection:
            try:
                glue.delete_connection(
                    CatalogId=assume_account,
                    ConnectionName=jdbc_conn.glue_connection
                )
            except Exception as e:
                logger.error(traceback.format_exc())
                err.append(str(e))
        crud.delete_jdbc_connection(provider_id, account, region, instance_id)
        try:
            crud.update_jdbc_instance_count(provider_id, account, region)
        except Exception as e:
            logger.error(traceback.format_exc())
            err.append(str(e))

    if not err:
        logger.error(err)
        # raise BizException(MessageEnum.SOURCE_S3_CONNECTION_DELETE_ERROR.get_code(), err)

    return True

def hide_jdbc_connection(provider_id: int, account: str, region: str, instance_id: str):
    database_type = convert_provider_id_2_database_type(provider_id)
    before_delete_jdbc_connection(provider_id, account, region, instance_id, database_type)

    err = []
    crud.hide_jdbc_connection(provider_id, account, region, instance_id)
    try:
        crud.update_jdbc_instance_count(provider_id, account, region)
    except Exception as e:
        logger.error(traceback.format_exc())
        err.append(str(e))

    if not err:
        logger.error(err)

    return True

def gen_credentials(account: str):
    try:
        iam_role_name = crud.get_iam_role(account)
        assumed_role = sts.assume_role(
            RoleArn=f"{iam_role_name}",
            RoleSessionName="glue-rds-connection"
        )
        credentials = assumed_role['Credentials']
    except Exception as err:
        logger.error(err)
        raise BizException(MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_code(),
                           MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_msg())
    if credentials is None:
        raise BizException(MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_code(),
                           MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_msg())
    return credentials

# Create crawler, connection and database and start crawler
def sync_rds_connection(account: str, region: str, instance_name: str, rds_user=None, rds_password=None,
                        rds_secret_id=None):
    glue_connection_name = f"{const.SOLUTION_NAME}-{DatabaseType.RDS.value}-{instance_name}"
    glue_database_name = f"{const.SOLUTION_NAME}-{DatabaseType.RDS.value}-{instance_name}"
    crawler_name = f"{const.SOLUTION_NAME}-{DatabaseType.RDS.value}-{instance_name}"
    if rds_user is None and rds_secret_id is None:
        raise BizException(MessageEnum.SOURCE_RDS_NO_AUTH.get_code(),
                           MessageEnum.SOURCE_RDS_NO_AUTH.get_msg())

    if rds_user is not None and rds_secret_id is not None:
        raise BizException(MessageEnum.SOURCE_RDS_DUPLICATE_AUTH.get_code(),
                           MessageEnum.SOURCE_RDS_DUPLICATE_AUTH.get_msg())

    if region != admin_region:
        raise BizException(MessageEnum.SOURCE_DO_NOT_SUPPORT_CROSS_REGION.get_code(),
                           MessageEnum.SOURCE_DO_NOT_SUPPORT_CROSS_REGION.get_msg())

    state = crud.get_rds_instance_source_glue_state(account, region, instance_name)
    logger.info("sync_rds_connection state is:")
    logger.info(state)
    if state == ConnectionState.PENDING.value:
        raise BizException(MessageEnum.SOURCE_CONNECTION_NOT_FINISHED.get_code(),
                           MessageEnum.SOURCE_CONNECTION_NOT_FINISHED.get_msg())

    elif state == ConnectionState.CRAWLING.value:
        raise BizException(MessageEnum.SOURCE_CONNECTION_CRAWLING.get_code(),
                           MessageEnum.SOURCE_CONNECTION_CRAWLING.get_msg())

    crawler_role_arn = __gen_role_arn(account_id=account,
                                      region=region,
                                      role_name='GlueDetectionJobRole')
    credentials = gen_credentials(account)
    try:
        # PENDING｜ACTIVE｜ERROR with message
        crud.set_rds_instance_source_glue_state(account, region, instance_name, ConnectionState.PENDING.value)
        rds = boto3.client(
            'rds',
            aws_access_key_id=credentials['AccessKeyId'],
            aws_secret_access_key=credentials['SecretAccessKey'],
            aws_session_token=credentials['SessionToken'],
            region_name=region
        )
        """ :type : pyboto3.rds """

        rds_instance = rds.describe_db_instances(DBInstanceIdentifier=instance_name)['DBInstances'][0]
        logger.info("sync_rds_connection describe_db_instances ")
        logger.info(rds_instance)
        engine = rds_instance['Engine']
        host = rds_instance['Endpoint']['Address']
        port = rds_instance['Endpoint']['Port']
        rds_vpc_id = rds_instance['DBSubnetGroup']['VpcId']
        rds_az = rds_instance['AvailabilityZone']
        # Currently Glue crawler connection does not support cn-north-1d.
        if rds_az == 'cn-north-1d':
            rds_az = 'cn-north-1a'
        rds_security_groups = []
        for vpc_sg in rds_instance['VpcSecurityGroups']:
            if vpc_sg['Status'] == 'active':
                rds_security_groups.append(vpc_sg['VpcSecurityGroupId'])

        ec2_client = boto3.client('ec2',
                                  aws_access_key_id=credentials['AccessKeyId'],
                                  aws_secret_access_key=credentials['SecretAccessKey'],
                                  aws_session_token=credentials['SessionToken'],
                                  region_name=region)
        response = ec2_client.describe_subnets(
            Filters=[
                {
                    'Name': 'vpc-id',
                    'Values': [rds_vpc_id]
                },
                {
                    'Name': 'availability-zone',
                    'Values': [rds_az]
                }
            ]
        )
        logger.debug(response)
        # Private subnet priority
        rds_subnet_id = const.EMPTY_STR
        public_subnet_id = const.EMPTY_STR
        has_outbound_route = False
        for subnet_desc in response['Subnets']:
            subnet_map_public_ip = subnet_desc['MapPublicIpOnLaunch']
            if subnet_map_public_ip:
                logger.info(f"subnet({subnet_desc['SubnetId']}) is public")
                public_subnet_id = subnet_desc['SubnetId']
            else:
                logger.info(f"subnet({subnet_desc['SubnetId']}) is private")
                rds_subnet_id = subnet_desc['SubnetId']
                has_outbound_route = check_subnet_has_outbound_route(credentials, region, subnet_desc['SubnetId'])
                if has_outbound_route:
                    break

        if rds_subnet_id == const.EMPTY_STR:
            if public_subnet_id == const.EMPTY_STR:
                raise BizException(MessageEnum.SOURCE_RDS_NO_PRIVATE_ACCESSABLE.get_code(),
                                   MessageEnum.SOURCE_RDS_NO_PRIVATE_ACCESSABLE.get_msg())
            else:
                rds_subnet_id = public_subnet_id

        if not has_outbound_route:
            check_link(credentials, region, rds_vpc_id, rds_secret_id)
        if rds_secret_id is not None:
            secretsmanager = boto3.client('secretsmanager',
                                          aws_access_key_id=credentials['AccessKeyId'],
                                          aws_secret_access_key=credentials['SecretAccessKey'],
                                          aws_session_token=credentials['SessionToken'],
                                          region_name=region
                                          )
            """ :type : pyboto3.secretsmanager """
            secret_value = secretsmanager.get_secret_value(
                SecretId=rds_secret_id
            )
            secret_values = json.loads(secret_value['SecretString'])
            rds_user = secret_values['username']
            rds_password = secret_values['password']

        payload = {
            "engine": engine,
            "host": host,
            "port": port,
            "username": rds_user,
            "password": rds_password,
        }
        schema_list = __list_rds_schema(account, region, credentials, instance_name, payload, rds_security_groups,
                                        rds_subnet_id)
        logger.info("sync_rds_connection schema_list :")
        logger.info(schema_list)
        if len(schema_list) > 0:
            glue = boto3.client('glue',
                                aws_access_key_id=credentials['AccessKeyId'],
                                aws_secret_access_key=credentials['SecretAccessKey'],
                                aws_session_token=credentials['SessionToken'],
                                region_name=region
                                )
            """ :type : pyboto3.glue """

            jdbc_url = __create_jdbc_url(engine=engine, host=host, port=port)
            try:
                glue.get_connection(Name=glue_connection_name)
            except Exception as e:
                logger.info("sync_rds_connection get_connection error and create:")
                logger.info(str(e))
                if rds_secret_id is None:
                    response = glue.create_connection(
                        ConnectionInput={
                            'Name': glue_connection_name,
                            'Description': glue_connection_name,
                            'ConnectionType': 'JDBC',
                            'ConnectionProperties': {
                                'USERNAME': rds_user,
                                'PASSWORD': rds_password,
                                'JDBC_CONNECTION_URL': jdbc_url,
                                'JDBC_ENFORCE_SSL': 'false',
                            },

                            'PhysicalConnectionRequirements': {
                                'SubnetId': rds_subnet_id,
                                'AvailabilityZone': rds_az,
                                'SecurityGroupIdList': rds_security_groups
                            }
                        },
                        Tags={
                            const.TAG_KEY: const.TAG_VALUE,
                            const.TAG_ADMIN_ACCOUNT_ID: admin_account_id
                        }
                    )
                    logger.info(response)
                else:
                    response = glue.create_connection(
                        ConnectionInput={
                            'Name': glue_connection_name,
                            'Description': glue_connection_name,
                            'ConnectionType': 'JDBC',
                            'ConnectionProperties': {
                                'SECRET_ID': rds_secret_id,
                                'JDBC_CONNECTION_URL': jdbc_url,
                                'JDBC_ENFORCE_SSL': 'false',
                            },

                            'PhysicalConnectionRequirements': {
                                'SubnetId': rds_subnet_id,
                                'AvailabilityZone': rds_az,
                                'SecurityGroupIdList': rds_security_groups
                            }
                        },
                        Tags={
                            const.TAG_KEY: const.TAG_VALUE,
                            const.TAG_ADMIN_ACCOUNT_ID: admin_account_id
                        },
                    )
                    logger.info(response)
            try:
                glue.get_database(Name=glue_database_name)
            except Exception as e:
                logger.info("sync_rds_connection get_database error and create:")
                logger.info(str(e))
                response = glue.create_database(DatabaseInput={'Name': glue_database_name})
                logger.info(response)
            lakeformation = boto3.client('lakeformation',
                                         aws_access_key_id=credentials['AccessKeyId'],
                                         aws_secret_access_key=credentials['SecretAccessKey'],
                                         aws_session_token=credentials['SessionToken'],
                                         region_name=region)
            """ :type : pyboto3.lakeformation """
            # retry for grant permissions
            num_retries = GRANT_PERMISSIONS_RETRIES
            while num_retries > 0:
                try:
                    response = lakeformation.grant_permissions(
                        Principal={
                            'DataLakePrincipalIdentifier': f"{crawler_role_arn}"
                        },
                        Resource={
                            'Database': {
                                'Name': glue_database_name
                            }
                        },
                        Permissions=['ALL'],
                        PermissionsWithGrantOption=['ALL']
                    )
                except Exception as e:
                    sleep(SLEEP_MIN_TIME)
                    num_retries -= 1
                else:
                    break
            else:
                raise BizException(MessageEnum.SOURCE_UNCONNECTED.get_code(), MessageEnum.SOURCE_UNCONNECTED.get_msg())
            jdbc_targets = []
            for schema in schema_list:
                jdbc_targets.append(
                    {
                        'ConnectionName': glue_connection_name,
                        'Path': f"{schema}/%",
                    }
                )
            logger.info("sync_rds_connection jdbc_targets:")
            logger.info(jdbc_targets)
            try:
                response = glue.get_crawler(Name=crawler_name)
                logger.info("sync_rds_connection get_crawler:")
                logger.info(response)
                try:
                    if not state or state == ConnectionState.ACTIVE.value or state == ConnectionState.UNSUPPORTED.value \
                            or state == ConnectionState.ERROR.value or state == ConnectionState.STOPPING.value:
                        up_cr_response = glue.update_crawler(
                            Name=crawler_name,
                            Role=crawler_role_arn,
                            DatabaseName=glue_database_name,
                            Targets={
                                'JdbcTargets': jdbc_targets,
                            },
                            SchemaChangePolicy={
                                'UpdateBehavior': 'UPDATE_IN_DATABASE',
                                'DeleteBehavior': 'DELETE_FROM_DATABASE'
                            }
                        )
                        logger.info("update rds crawler:")
                        logger.info(up_cr_response)
                except Exception as e:
                    logger.info("update_crawler error")
                    logger.info(str(e))
                st_cr_response = glue.start_crawler(
                    Name=crawler_name
                )
                logger.info(st_cr_response)
            except Exception as e:
                logger.info("sync_rds_connection get_crawler and create:")
                logger.info(str(e))
                response = glue.create_crawler(
                    Name=crawler_name,
                    Role=crawler_role_arn,
                    DatabaseName=glue_database_name,
                    Targets={
                        'JdbcTargets': jdbc_targets,
                    },
                    Tags={
                        const.TAG_KEY: const.TAG_VALUE,
                        const.TAG_ADMIN_ACCOUNT_ID: admin_account_id
                    },
                )
                logger.info(response)
                start_response = glue.start_crawler(
                    Name=crawler_name
                )
                logger.info(start_response)
            crud.create_rds_connection(account, region, instance_name, glue_connection_name, glue_database_name,
                                       None, crawler_name)
        else:
            crud.set_rds_instance_source_glue_state(account, region, instance_name,
                                                    MessageEnum.SOURCE_RDS_NO_SCHEMA.get_msg())
            raise BizException(MessageEnum.SOURCE_RDS_NO_SCHEMA.get_code(), MessageEnum.SOURCE_RDS_NO_SCHEMA.get_msg())
    except Exception as err:
        crud.set_rds_instance_source_glue_state(account, region, instance_name, str(err))
        glue = boto3.client('glue',
                            aws_access_key_id=credentials['AccessKeyId'],
                            aws_secret_access_key=credentials['SecretAccessKey'],
                            aws_session_token=credentials['SessionToken'],
                            region_name=region
                            )
        """ :type : pyboto3.glue """

        try:
            glue.delete_crawler(crawler_name)
        except Exception:
            pass
        try:
            glue.delete_database(Name=glue_database_name)
        except Exception:
            pass
        try:
            glue.delete_connection(ConnectionName=glue_connection_name)
        except Exception:
            pass

        logger.error(traceback.format_exc())
        raise BizException(MessageEnum.SOURCE_CONNECTION_FAILED.get_code(),
                           str(err))


# raise error if cannot delete data source for rds
def before_delete_rds_connection(account: str, region: str, instance: str):
    rds_instance = crud.get_rds_instance_source(account, region, instance)
    if rds_instance is None:
        raise BizException(MessageEnum.SOURCE_RDS_NO_INSTANCE.get_code(),
                           MessageEnum.SOURCE_RDS_NO_INSTANCE.get_msg())
    # crawler, if crawling try to stop and raise, if pending raise directly
    state = crud.get_rds_instance_source_glue_state(account, region, instance)
    if state == ConnectionState.PENDING.value:
        raise BizException(MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_code(),
                           MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_msg())
    elif state == ConnectionState.CRAWLING.value:
        try:
            # Stop the crawler
            __glue(account=account, region=region).stop_crawler(Name=rds_instance.glue_crawler)
        except Exception as e:
            logger.error(traceback.format_exc())
        raise BizException(MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_code(),
                           MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_msg())
    elif state == ConnectionState.ACTIVE.value:
        # if job running, do not stop but raising
        if not can_delete_job_database(account_id=account, region=region, database_type='rds',
                                       database_name=instance):
            raise BizException(MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_code(),
                               MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_msg())


def delete_rds_connection(account: str, region: str, instance: str, delete_catalog_only=False):
    before_delete_rds_connection(account, region, instance)
    glue = __glue(account, region)
    rds_instance = crud.get_rds_instance_source(account, region, instance)
    err = []
    # 1/3 delete job database
    try:
        if not delete_catalog_only:
            delete_job_database(account_id=account, region=region, database_type='rds', database_name=instance)
    except Exception as e:
        err.append(str(e))
    # 2/3 delete catalog
    try:
        delete_catalog_by_database_region(database=instance, region=region, type='rds')
    except Exception as e:
        err.append(str(e))

    if not delete_catalog_only:
        # 3/3 delete source
        try:
            glue.delete_crawler(Name=rds_instance.glue_crawler)
        except Exception as e:
            err.append(e)
        try:
            glue.delete_database(Name=rds_instance.glue_database)
        except Exception as e:
            err.append(e)
        try:
            glue.delete_connection(ConnectionName=rds_instance.glue_connection)
        except Exception as e:
            err.append(e)
        crud.delete_rds_connection(account, region, instance)
        if err:
            logger.error(traceback.format_exc())
            # raise BizException(MessageEnum.SOURCE_RDS_CONNECTION_DELETE_ERROR.get_code(), f"Delete with error: {err}")

    return []

def hide_rds_connection(account: str, region: str, instance: str):
    before_delete_rds_connection(account, region, instance)
    crud.hide_rds_connection(account, region, instance)
    return []

def list_s3_bucket_source(condition: QueryCondition):
    return crud.list_s3_bucket_source(condition)


# raise error if cannot delete data source for s3
def before_delete_s3_connection(account: str, region: str, bucket: str):
    s3_bucket = crud.get_s3_bucket_source(account, region, bucket)
    if s3_bucket is None:
        raise BizException(MessageEnum.SOURCE_S3_NO_BUCKET.get_code(),
                           MessageEnum.SOURCE_S3_NO_BUCKET.get_msg())
    state = crud.get_s3_bucket_source_glue_state(account, region, bucket)
    if state == ConnectionState.PENDING.value:
        raise BizException(MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_code(),
                           MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_msg())
    elif state == ConnectionState.CRAWLING.value:
        try:
            # Stop the crawler
            __glue(account=account, region=region).stop_crawler(Name=s3_bucket.glue_crawler)
        except Exception as e:
            logger.error(traceback.format_exc())
        raise BizException(MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_code(),
                           MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_msg())
    elif state == ConnectionState.ACTIVE.value:
        # if job running, do not stop but raising
        if not can_delete_job_database(account_id=account, region=region, database_type='s3',
                                       database_name=bucket):
            raise BizException(MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_code(),
                               MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_msg())


def delete_s3_connection(account: str, region: str, bucket: str, delete_catalog_only=False):
    before_delete_s3_connection(account, region, bucket)
    err = []
    # 1/3 delete job database
    try:
        if not delete_catalog_only:
            delete_job_database(account_id=account, region=region, database_type='s3', database_name=bucket)
    except Exception as e:
        err.append(str(e))
    # 2/3 delete catalog
    try:
        delete_catalog_by_database_region(database=bucket, region=region, type='s3')
    except Exception as e:
        err.append(str(e))
    # 3/3 delete source
    if not delete_catalog_only:
        s3_bucket = crud.get_s3_bucket_source(account, region, bucket)
        glue = __glue(account=account, region=region)
        try:
            glue.delete_crawler(Name=s3_bucket.glue_crawler)
        except Exception as e:
            err.append(str(e))
        try:
            glue.delete_database(Name=s3_bucket.glue_database)
        except Exception as e:
            err.append(str(e))

        crud.delete_s3_bucket_connection(account, region, bucket)
        try:
            crud.update_s3_bucket_count(account, region)
        except Exception as e:
            err.append(str(e))

    if err:
        logger.error(traceback.format_exc())
        # raise BizException(MessageEnum.SOURCE_S3_CONNECTION_DELETE_ERROR.get_code(), err)

    return True

def hide_s3_connection(account: str, region: str, bucket: str):
    before_delete_s3_connection(account, region, bucket)
    err = []

    crud.hide_s3_bucket_connection(account, region, bucket)
    try:
        crud.update_s3_bucket_count(account, region)
    except Exception as e:
        err.append(str(e))

    if err:
        logger.error(traceback.format_exc())
        # raise BizException(MessageEnum.SOURCE_S3_CONNECTION_DELETE_ERROR.get_code(), err)

    return True

def delete_glue_connection(account: str, region: str, glue_crawler: str,
                           glue_database: str, glue_connection: str):
    glue = __glue(account=account, region=region)
    try:
        glue.delete_crawler(Name=glue_crawler)
    except Exception as e:
        logger.info("delete_glue_crawler" + str(e))
    try:
        glue.delete_database(Name=glue_database)
    except Exception as e:
        logger.info("delete_glue_database" + str(e))
    # try:
    #     glue.delete_connection(ConnectionName=glue_connection)
    # except Exception as e:
    #     logger.info("delete_glue_connection" + str(e))
    return True


def refresh_data_source(provider_id: int, accounts: list[str], type: str):
    # tmp_provider = int(provider)
    # identify the real provider id for the jdbc proxy type
    for account_id in accounts:
        account = crud.get_account_by_id(account_id=account_id)
        if account:
            provider_id = account.account_provider_id
            if provider_id == Provider.AWS_CLOUD.value:
                refresh_aws_data_source([account_id], type)
            else:
                refresh_third_data_source(provider_id, [account_id], type)


def refresh_aws_data_source(accounts: list[str], type: str):
    if type is None or len(accounts) == 0:
        raise BizException(MessageEnum.SOURCE_REFRESH_FAILED.get_code(),
                           MessageEnum.SOURCE_REFRESH_FAILED.get_msg())
    try:
        if type == DataSourceType.s3.value:
            s3_detector.detect(accounts)

        elif type == DataSourceType.rds.value:
            rds_detector.detect(accounts)

        elif type == DataSourceType.glue_database.value:
            glue_database_detector.detect(accounts)

        elif type == DataSourceType.jdbc.value:
            jdbc_detector.detect(Provider.AWS_CLOUD.value, accounts)
        elif type == DataSourceType.all.value:
            s3_detector.detect(accounts)
            rds_detector.detect(accounts)
            glue_database_detector.detect(accounts)
            jdbc_detector.detect(Provider.AWS_CLOUD.value, accounts)
    except Exception as e:
        logger.error(traceback.format_exc())
        raise BizException(MessageEnum.SOURCE_CONNECTION_FAILED.get_code(), str(e))

def refresh_third_data_source(provider_id: int, accounts: list[str], type: str):
    if type is None or not accounts:
        raise BizException(MessageEnum.SOURCE_REFRESH_FAILED.get_code(),
                           MessageEnum.SOURCE_REFRESH_FAILED.get_msg())
    try:
        jdbc_detector.detect(provider_id, accounts)
    except Exception as e:
        logger.error(traceback.format_exc())
        raise BizException(MessageEnum.SOURCE_CONNECTION_FAILED.get_code(), str(e))

def get_data_source_coverage(provider_id):
    provider_id = int(provider_id)
    if provider_id == Provider.AWS_CLOUD.value:
        res = SourceCoverage(
            s3_total=crud.get_total_s3_buckets_count(),
            s3_connected=crud.get_connected_s3_buckets_size(),
            rds_total=crud.get_total_rds_instances_count(),
            rds_connected=crud.get_connected_rds_instances_count(),
            glue_total=crud.get_total_glue_database_count(),
            glue_connected=crud.get_connected_glue_database_count(),
            jdbc_total=crud.get_total_jdbc_instances_count(provider_id) + crud.get_total_jdbc_instances_count(
                Provider.JDBC_PROXY.value),
            jdbc_connected=crud.get_connected_jdbc_instances_count(
                provider_id) + crud.get_connected_jdbc_instances_count(Provider.JDBC_PROXY.value)
        )
    else:
        res = SourceCoverage(
            jdbc_total=crud.get_total_jdbc_instances_count(provider_id),
            jdbc_connected=crud.get_connected_jdbc_instances_count(provider_id)
        )
    return res


# Update account list by stackset state
# It will clean up accounts which stack_id is not null
def reload_organization_account(it_account: str):
    stack_instances = []
    member_accounts = []
    delegated_role_arn = __gen_role_arn(
        account_id=it_account,
        region=admin_region,
        role_name=_delegated_role_name)
    if __assume_role(it_account, delegated_role_arn):
        # do not clean up for autosync
        # crud.cleanup_delegated_account(delegated_account_id=it_account)
        assumed_role = sts.assume_role(
            RoleArn=delegated_role_arn,
            RoleSessionName='get_organization'
        )
        credentials = assumed_role['Credentials']
        try:
            # check if it is delegated admin
            orgs = boto3.client('organizations',
                                aws_access_key_id=credentials['AccessKeyId'],
                                aws_secret_access_key=credentials['SecretAccessKey'],
                                aws_session_token=credentials['SessionToken'],
                                region_name=admin_region
                                )
            delegated = orgs.list_delegated_administrators()
            call_as = 'DELEGATED_ADMIN' if any(
                admin['Id'] == it_account for admin in delegated['DelegatedAdministrators']) else 'SELF'
            cloudformation = boto3.client('cloudformation',
                                          aws_access_key_id=credentials['AccessKeyId'],
                                          aws_secret_access_key=credentials['SecretAccessKey'],
                                          aws_session_token=credentials['SessionToken'],
                                          region_name=admin_region
                                          )
            active_stack_sets = cloudformation.list_stack_sets(
                Status='ACTIVE',
                CallAs=call_as
            )['Summaries']
            for stack_set in active_stack_sets:
                stack = cloudformation.describe_stack_set(
                    StackSetName=stack_set['StackSetName'],
                    CallAs=call_as
                )
                if '(SO8031-sub)' in stack['StackSet']['TemplateBody']:
                    response = cloudformation.list_stack_instances(
                        StackSetName=stack_set['StackSetName'],
                        MaxResults=30,
                        CallAs=call_as
                    )
                    stack_instances.extend(response['Summaries'])

                    while 'NextToken' in response:
                        response = cloudformation.list_stack_instances(
                            StackSetName=stack_set['StackSetName'],
                            MaxResults=30,
                            NextToken=response['NextToken'],
                            CallAs=call_as
                        )
                        stack_instances.extend(response['Summaries'])
            logger.debug(stack_instances)
            # convert to unique account list
            member_accounts = list(set(item['Account'] for item in stack_instances))
            for account in member_accounts:
                add_aws_account(account)
        except Exception as e:
            logger.error(traceback.format_exc())
            raise BizException(MessageEnum.SOURCE_ORG_ADD_ACCOUNT_FAILED.get_code(), str(e))
    else:
        raise BizException(MessageEnum.SOURCE_ASSUME_DELEGATED_ROLE_FAILED.get_code(),
                           MessageEnum.SOURCE_ASSUME_DELEGATED_ROLE_FAILED.get_msg())
    return member_accounts


def add_account(account: SourceNewAccount):
    if crud.get_account_by_id(account_id=account.account_id):
        raise BizException(MessageEnum.SOURCE_ACCOUNT_ID_ALREADY_EXISTS.get_code(),
                           MessageEnum.SOURCE_ACCOUNT_ID_ALREADY_EXISTS.get_msg())

    if account.account_provider == Provider.AWS_CLOUD.value:
        add_aws_account(account.account_id)
    else:
        add_third_account(account, admin_account_id, admin_region)


def add_aws_account(account_id: str):
    # Open this loop for multiple region support
    # for region in const.CN_REGIONS:
    assumed_role = False
    for region in [admin_region]:
        role_arn = __gen_role_arn(account_id=account_id, region=region, role_name=_agent_role_name)
        if __assume_role(account_id, role_arn):
            assumed_role = True
            # raise BizException(MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_code(),
            #  MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_msg())
            crud.add_account(
                aws_account_id=account_id,
                aws_account_alias=None,
                aws_account_email=None,
                region=region,
                status=1,
                detection_role_status='SUCCESS',
                delegated_aws_account_id=None,
                organization_unit_id=None,
                stack_id=None,
                stackset_id=None,
                stackset_name=None,
                stack_status=None,
                stack_instance_status=None,
                detection_role_name=role_arn
            )
            __update_access_policy_for_account()
            # add account and retrieve data source for all
            refresh_data_source(provider_id=Provider.AWS_CLOUD.value, accounts=[account_id], type='all')
        else:
            logger.debug(MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_msg())
    if not assumed_role:
        raise BizException(MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_code(),
                           MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_msg())

def add_third_account(account, admin_account, admin_region):
    role_arn = __gen_role_arn(account_id=admin_account, region=admin_region, role_name=_agent_role_name)
    crud.add_third_account(account, role_arn)

def delete_account(account_provider: int, account_id: str, region: str):
    account = crud.get_account_by_id(account_id=account_id)
    if account:
        account_provider = account.account_provider_id
        if account_provider == Provider.AWS_CLOUD.value:
            delete_aws_account(account_id)
        else:
            delete_third_account(account_provider, account_id, region)

def delete_aws_account(account_id):
    accounts_by_region = crud.list_all_accounts_by_region(region=admin_region)
    list_accounts = [c[0] for c in accounts_by_region]
    if account_id not in list_accounts:
        raise BizException(MessageEnum.SOURCE_ACCOUNT_NOT_EXIST.get_code(),
                           MessageEnum.SOURCE_ACCOUNT_NOT_EXIST.get_msg())
    # this loop for multiple region support
    # for region in const.CN_REGIONS:
    assumed_role = False
    for region in [admin_region]:
        role_arn = __gen_role_arn(account_id=account_id, region=region, role_name=_agent_role_name)
        if __assume_role(account_id, role_arn):
            assumed_role = True
        else:
            logger.debug(MessageEnum.SOURCE_ASSUME_ROLE_FAILED.get_msg())
    if not assumed_role:
        # account agent is undeployed, delete permanently, only infra
        del_error = False
        try:
            # delete jobs in db
            delete_job_by_account(account_id=account_id, region=admin_region)
        except Exception:
            del_error = True
            logger.error(traceback.format_exc())
        try:
            # delete catalogs in db
            delete_catalog_by_account(account_id=account_id, region=admin_region)
        except Exception:
            del_error = True
            logger.error(traceback.format_exc())
        try:
            # delete data sources in db
            __delete_data_source_by_account(account_id=account_id, region=admin_region)
        except Exception:
            del_error = True
            logger.error(traceback.format_exc())
        try:
            # delete account in db
            __delete_account(account_id=account_id, region=admin_region)
        except Exception:
            del_error = True
            logger.error(traceback.format_exc())
        try:
            # update policy
            __update_access_policy_for_account()
        except Exception:
            del_error = True
            logger.error(traceback.format_exc())
        if del_error:
            raise BizException(MessageEnum.SOURCE_ACCOUNT_DELETE_FAILED.get_code(),
                               MessageEnum.SOURCE_ACCOUNT_DELETE_FAILED.get_msg())
    else:
        # account agent exists, need uninsatll agent first
        raise BizException(MessageEnum.SOURCE_ACCOUNT_AGENT_EXIST.get_code(),
                           MessageEnum.SOURCE_ACCOUNT_AGENT_EXIST.get_msg())


def delete_third_account(account_provider, account_id, region):
    crud.delete_third_account(account_provider, account_id, region)
    try:
        delete_catalog_by_account(account_id=account_id, region=region)
    except Exception:
        del_error = True
        logger.error(traceback.format_exc())


def refresh_account():
    __update_access_policy_for_account()


def get_secrets(provider: int, account: str, region: str):
    account_id = account if provider == Provider.AWS_CLOUD.value else admin_account_id
    region_aws = region if provider == Provider.AWS_CLOUD.value else admin_region
    iam_role_name = crud.get_iam_role(account_id)

    assumed_role = sts.assume_role(
        RoleArn=f"{iam_role_name}",
        RoleSessionName="glue-s3-connection"
    )
    credentials = assumed_role['Credentials']
    secretsmanager = boto3.client('secretsmanager',
                                  aws_access_key_id=credentials['AccessKeyId'],
                                  aws_secret_access_key=credentials['SecretAccessKey'],
                                  aws_session_token=credentials['SessionToken'],
                                  region_name=region_aws
                                  )
    """ :type : pyboto3.secretsmanager """
    response = secretsmanager.list_secrets()
    secrets = []
    for secret in response['SecretList']:
        secrets.append(
            {
                "Name": secret['Name'],
                "ARN": secret['ARN']
            }
        )
    return secrets


def get_admin_account_info():
    return AdminAccountInfo(account_id=admin_account_id, region=admin_region)

def import_glue_database(glueDataBase: SourceGlueDatabaseBase):
    list = crud.list_glue_database_by_account(glueDataBase.account_id, glueDataBase.region, glueDataBase.glue_database_name)
    if list:
        raise BizException(MessageEnum.SOURCE_GLUE_DATABASE_EXISTS.get_code(),
                           MessageEnum.SOURCE_GLUE_DATABASE_EXISTS.get_msg())
    response = __glue(account=glueDataBase.account_id, region=glueDataBase.region).get_database(CatalogId=glueDataBase.account_id,
                                                                                                Name=glueDataBase.glue_database_name)['Database']
    # return response
    crud.import_glue_database(glueDataBase, response)

def update_jdbc_conn(jdbc_conn: JDBCInstanceSource):
    jdbc_source = JdbcSource(connection_url=jdbc_conn.jdbc_connection_url,
                             username=jdbc_conn.master_username,
                             password=jdbc_conn.password,
                             secret_id=jdbc_conn.secret)
    dbnames = get_db_names_4_jdbc(jdbc_source, jdbc_conn.jdbc_connection_schema)
    account_id, region = __get_admin_info(jdbc_conn)
    res: JDBCInstanceSourceFullInfo = crud.get_jdbc_instance_source_glue(jdbc_conn.account_provider_id,
                                                                         jdbc_conn.account_id,
                                                                         jdbc_conn.region,
                                                                         jdbc_conn.instance_id)
    check_connection(res, jdbc_conn, account_id, region)
    update_connection(res, jdbc_conn, account_id, region, dbnames)

def check_connection(res: JDBCInstanceSourceFullInfo, jdbc_instance: JDBCInstanceSource, assume_account, assume_role):
    if not res:
        raise BizException(MessageEnum.SOURCE_JDBC_CONNECTION_NOT_EXIST.get_code(),
                           MessageEnum.SOURCE_JDBC_CONNECTION_NOT_EXIST.get_msg())
    if res.glue_state == ConnectionState.PENDING.value:
        raise BizException(MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_code(),
                           MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_msg())
    elif res.glue_state == ConnectionState.CRAWLING.value:
        try:
            # Stop the crawler
            __glue(account=assume_account, region=assume_role).stop_crawler(Name=res.glue_crawler)
        except Exception as e:
            logger.error(traceback.format_exc())
        raise BizException(MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_code(),
                           MessageEnum.SOURCE_DELETE_WHEN_CONNECTING.get_msg())
    elif res.glue_state == ConnectionState.ACTIVE.value:
        # if job running, do not stop but raising
        if not can_delete_job_database(account_id=jdbc_instance.account_id, region=jdbc_instance.region, database_type='jdbc',
                                       database_name=jdbc_instance):
            raise BizException(MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_code(),
                               MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_msg())
    else:
        pass

def update_connection(res: JDBCInstanceSourceFullInfo, jdbc_instance: JDBCInstanceSourceUpdate, assume_account, assume_region, db_names):
    jdbc_targets = __gen_jdbc_targets_from_db_names(res.glue_connection, db_names)
    connectionProperties_dict = gen_conn_properties(jdbc_instance)
    __glue(account=assume_account, region=assume_region).update_connection(
        CatalogId=assume_account,
        Name=res.glue_connection,
        ConnectionInput={
            'Name': res.glue_connection,
            'Description': jdbc_instance.description,
            'ConnectionType': 'JDBC',
            'ConnectionProperties': connectionProperties_dict,
            'PhysicalConnectionRequirements': {
                'SubnetId': jdbc_instance.network_subnet_id,
                'SecurityGroupIdList': [
                    jdbc_instance.network_sg_id,
                ],
                'AvailabilityZone': jdbc_instance.network_availability_zone
            }
        }
    )
    crawler_role_arn = __gen_role_arn(account_id=assume_account,
                                      region=assume_region,
                                      role_name='GlueDetectionJobRole')
    # Update Crawler
    __update_crawler(res.account_provider_id,
                     res.account_id,
                     res.instance_id,
                     res.region,
                     jdbc_targets,
                     res.glue_crawler,
                     res.glue_database,
                     crawler_role_arn)
    crud.update_jdbc_connection_full(jdbc_instance)

def __validate_jdbc_url(url: str):
    for pattern in _jdbc_url_patterns:
        if re.match(pattern, url):
            return True

def add_jdbc_conn(jdbcConn: JDBCInstanceSource):
    jdbc_targets = []
    create_connection_response = {}
    # get_db_names(jdbcConn.jdbc_connection_url, jdbcConn.jdbc_connection_schema)
    account_id, region = __get_admin_info(jdbcConn)
    crawler_role_arn = __gen_role_arn(account_id=account_id,
                                      region=region,
                                      role_name='GlueDetectionJobRole')
    list = crud.list_jdbc_instance_source_by_instance_id_account(jdbcConn, account_id)
    if list:
        raise BizException(MessageEnum.SOURCE_JDBC_ALREADY_EXISTS.get_code(),
                           MessageEnum.SOURCE_JDBC_ALREADY_EXISTS.get_msg())
    glue_connection_name, glue_database_name, crawler_name = __gen_resources_name(jdbcConn)
    ec2_client, __ = __ec2(account=account_id, region=region)
    glue = __get_glue_client(account=account_id, region=region)
    try:
        availability_zone = ec2_client.describe_subnets(SubnetIds=[jdbcConn.network_subnet_id])['Subnets'][0]['AvailabilityZone']
        try:
            connectionProperties_dict = gen_conn_properties(jdbcConn)
            create_connection_response = __glue(account=account_id, region=region).create_connection(
                CatalogId=account_id,
                ConnectionInput={
                    'Name': glue_connection_name,
                    'Description': jdbcConn.description,
                    'ConnectionType': 'JDBC',
                    'ConnectionProperties': connectionProperties_dict,
                    'PhysicalConnectionRequirements': {
                        'SubnetId': jdbcConn.network_subnet_id,
                        'SecurityGroupIdList': [
                            jdbcConn.network_sg_id
                        ],
                        'AvailabilityZone': availability_zone
                    }
                },
                Tags={
                    const.TAG_KEY: const.TAG_VALUE,
                    const.TAG_ADMIN_ACCOUNT_ID: admin_account_id
                },
            )
        except ClientError as ce:
            if ce.response['Error']['Code'] == 'InvalidInputException':
                raise BizException(MessageEnum.SOURCE_JDBC_INPUT_INVALID.get_code(),
                                   MessageEnum.SOURCE_JDBC_INPUT_INVALID.get_msg())
            if ce.response['Error']['Code'] == 'AlreadyExistsException':
                raise BizException(MessageEnum.SOURCE_JDBC_ALREADY_EXISTS.get_code(),
                                   MessageEnum.SOURCE_JDBC_ALREADY_EXISTS.get_msg())
        except Exception as e:
            logger.error(traceback.format_exc())
        if create_connection_response.get('ResponseMetadata', {}).get('HTTPStatusCode') != 200:
            raise BizException(MessageEnum.SOURCE_JDBC_CREATE_FAIL.get_code(),
                               MessageEnum.SOURCE_JDBC_CREATE_FAIL.get_msg())
        # Creare Glue database
        glue.create_database(DatabaseInput={'Name': glue_database_name})
        # Create Crawler
        jdbc_source = JdbcSource(connection_url=jdbcConn.jdbc_connection_url, username=jdbcConn.master_username, password=jdbcConn.password, secret_id=jdbcConn.secret)
        db_names = get_db_names_4_jdbc(jdbc_source, jdbcConn.jdbc_connection_schema)
        for db_name in db_names:
            trimmed_db_name = db_name.strip()
            if trimmed_db_name:
                jdbc_targets.append({
                    'ConnectionName': glue_connection_name,
                    'Path': f"{trimmed_db_name}/%"
                })
        try:
            glue.create_crawler(
                Name=crawler_name,
                Role=crawler_role_arn,
                DatabaseName=glue_database_name,
                Targets={
                    'JdbcTargets': jdbc_targets,
                },
                Tags={
                    const.TAG_KEY: const.TAG_VALUE,
                    const.TAG_ADMIN_ACCOUNT_ID: admin_account_id
                },
            )
        except Exception:
            logger.error(traceback.format_exc())

        jdbcConn.network_availability_zone = availability_zone
        jdbcConn.create_type = JDBCCreateType.ADD.value
        jdbc_conn_insert = JDBCInstanceSourceFullInfo()
        jdbc_conn_insert.instance_id = jdbcConn.instance_id
        jdbc_conn_insert.account_provider_id = jdbcConn.account_provider_id
        jdbc_conn_insert.account_id = jdbcConn.account_id
        jdbc_conn_insert.region = jdbcConn.region
        jdbc_conn_insert.detection_history_id = 0
        jdbc_conn_insert.description = jdbcConn.description
        jdbc_conn_insert.jdbc_connection_url = jdbcConn.jdbc_connection_url
        jdbc_conn_insert.jdbc_connection_schema = jdbcConn.jdbc_connection_schema
        jdbc_conn_insert.jdbc_enforce_ssl = jdbcConn.jdbc_enforce_ssl
        jdbc_conn_insert.kafka_ssl_enabled = jdbcConn.kafka_ssl_enabled
        jdbc_conn_insert.master_username = jdbcConn.master_username
        jdbc_conn_insert.skip_custom_jdbc_cert_validation = jdbcConn.skip_custom_jdbc_cert_validation
        jdbc_conn_insert.custom_jdbc_cert = jdbcConn.custom_jdbc_cert
        jdbc_conn_insert.custom_jdbc_cert_string = jdbcConn.custom_jdbc_cert_string
        jdbc_conn_insert.network_availability_zone = jdbcConn.network_availability_zone
        jdbc_conn_insert.network_subnet_id = jdbcConn.network_subnet_id
        jdbc_conn_insert.network_sg_id = jdbcConn.network_sg_id
        jdbc_conn_insert.creation_time = jdbcConn.creation_time
        jdbc_conn_insert.last_updated_time = jdbcConn.last_updated_time
        jdbc_conn_insert.jdbc_driver_class_name = jdbcConn.jdbc_driver_class_name
        jdbc_conn_insert.jdbc_driver_jar_uri = jdbcConn.jdbc_driver_jar_uri
        jdbc_conn_insert.create_type = jdbcConn.create_type
        jdbc_conn_insert.glue_connection = glue_connection_name
        jdbc_conn_insert.glue_crawler = crawler_name
        jdbc_conn_insert.glue_database = glue_database_name
        crud.add_jdbc_conn(jdbc_conn_insert)
    except ClientError as ce:
        logger.error(traceback.format_exc())
        if ce.response['Error']['Code'] == 'AlreadyExistsException':
            raise BizException(MessageEnum.SOURCE_JDBC_ALREADY_EXISTS.get_code(),
                               MessageEnum.SOURCE_JDBC_ALREADY_EXISTS.get_msg())
        elif ce.response['Error']['Code'] == 'InvalidSubnetID.NotFound':
            raise BizException(MessageEnum.SOURCE_SUBNET_NOT_EXIST.get_code(),
                               MessageEnum.SOURCE_SUBNET_NOT_EXIST.get_msg())
        else:
            raise BizException(MessageEnum.BIZ_UNKNOWN_ERR.get_code(),
                               MessageEnum.BIZ_UNKNOWN_ERR.get_msg())

def gen_conn_properties(jdbcConn):
    connectionProperties_dict = {}
    if jdbcConn.jdbc_enforce_ssl != 'false' and jdbcConn.custom_jdbc_cert:
        connectionProperties_dict['CUSTOM_JDBC_CERT'] = jdbcConn.custom_jdbc_cert
    if jdbcConn.jdbc_enforce_ssl != 'false' and jdbcConn.custom_jdbc_cert_string:
        connectionProperties_dict['CUSTOM_JDBC_CERT_STRING'] = jdbcConn.custom_jdbc_cert_string
    if jdbcConn.jdbc_driver_class_name:
        connectionProperties_dict['JDBC_DRIVER_CLASS_NAME'] = jdbcConn.jdbc_driver_class_name
    if jdbcConn.jdbc_driver_jar_uri:
        connectionProperties_dict['JDBC_DRIVER_JAR_URI'] = jdbcConn.jdbc_driver_jar_uri
    if jdbcConn.master_username:
        connectionProperties_dict['USERNAME'] = jdbcConn.master_username
    if jdbcConn.password:
        connectionProperties_dict['PASSWORD'] = jdbcConn.password
    if jdbcConn.secret:
        connectionProperties_dict['SECRET_ID'] = jdbcConn.secret
    connectionProperties_dict['JDBC_CONNECTION_URL'] = jdbcConn.jdbc_connection_url
    connectionProperties_dict['JDBC_ENFORCE_SSL'] = jdbcConn.jdbc_enforce_ssl
    if jdbcConn.skip_custom_jdbc_cert_validation and jdbcConn.skip_custom_jdbc_cert_validation == 'true':
        connectionProperties_dict['SKIP_CUSTOM_JDBC_CERT_VALIDATION'] = jdbcConn.skip_custom_jdbc_cert_validation
    return connectionProperties_dict

def test_jdbc_conn(jdbc_conn_param: JDBCInstanceSourceBase):
    res = "FAIL"
    account_id, region = __get_admin_info(jdbc_conn_param)
    cursor = None
    connection = None
    # get connection name from sdp db
    source: JDBCInstanceSourceFullInfo = crud.get_jdbc_instance_source_glue(provider_id=jdbc_conn_param.account_provider_id,
                                                                            account=jdbc_conn_param.account_id,
                                                                            region=jdbc_conn_param.region,
                                                                            instance_id=jdbc_conn_param.instance_id)
    if not source or not source.glue_connection:
        raise BizException(MessageEnum.SOURCE_JDBC_CONNECTION_NOT_EXIST.get_code(),
                           MessageEnum.SOURCE_JDBC_CONNECTION_NOT_EXIST.get_msg())
    try:
        response = __glue(account_id, region).get_connection(Name=source.glue_connection)
        connection_properties = response['Connection']['ConnectionProperties']
        username = connection_properties['USERNAME']
        password = connection_properties['PASSWORD']
        jdbc_url = connection_properties['JDBC_CONNECTION_URL']
        connection = pymysql.connect(host=jdbc_url, user=username, password=password)
        cursor = connection.cursor()
        cursor.execute("SELECT 1")
        result = cursor.fetchone()
        if result[0] == 1:
            logger.info("Glue Connection is reachable.")
            res = "SUCCESS"
        else:
            logger.info("Glue Connection is not reachable.")
    except Exception as e:
        logger.error(traceback.format_exc())
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
        jdbc_conn = JDBCInstanceSourceUpdateBase(account_provider_id=jdbc_conn_param.account_provider_id,
                                                 account_id=jdbc_conn_param.account_id,
                                                 region=jdbc_conn_param.region,
                                                 instance_id=jdbc_conn_param.instance_id,
                                                 connection_status=res
                                                 )
        crud.set_jdbc_instance_connection_status(jdbc_conn)


def import_jdbc_conn(jdbc_conn: JDBCInstanceSourceBase):
    res_connection = None
    if crud.list_jdbc_connection_by_connection(jdbc_conn.instance_id):
        raise BizException(MessageEnum.SOURCE_JDBC_ALREADY_IMPORTED.get_code(),
                           MessageEnum.SOURCE_JDBC_ALREADY_IMPORTED.get_msg())        
    account_id, region = __get_admin_info()
    try:
        res_connection = __glue(account_id, region).get_connection(Name=jdbc_conn.instance_id)['Connection']
    except ClientError as ce:
        logger.error(traceback.format_exc())
        if ce.response['Error']['Code'] == 'EntityNotFoundException':
            raise BizException(MessageEnum.SOURCE_CONNECTION_NOT_FOUND.get_code(),
                               MessageEnum.SOURCE_CONNECTION_NOT_FOUND.get_msg())
        else:
            raise BizException(MessageEnum.BIZ_UNKNOWN_ERR.get_code(),
                               MessageEnum.BIZ_UNKNOWN_ERR.get_msg())
    except Exception as e:
        logger.error(traceback.format_exc())
        raise BizException(MessageEnum.BIZ_UNKNOWN_ERR.get_code(),
                           MessageEnum.BIZ_UNKNOWN_ERR.get_msg())
    logger.info(f"connection info:{res_connection}")
    jdbc_conn_insert = JDBCInstanceSourceFullInfo()
    jdbc_conn_insert.account_id = jdbc_conn.account_id
    jdbc_conn_insert.region = jdbc_conn.region
    jdbc_conn_insert.account_provider_id = jdbc_conn.account_provider_id
    jdbc_conn_insert.detection_history_id = 0
    jdbc_conn_insert.instance_id = jdbc_conn.instance_id
    jdbc_conn_insert.glue_connection = jdbc_conn.instance_id
    jdbc_conn_insert.create_type = JDBCCreateType.IMPORT.value
    jdbc_conn_insert.description = res_connection.get('Description')
    if res_connection.get('ConnectionProperties'):
        jdbc_conn_insert.jdbc_connection_url = res_connection.get('ConnectionProperties').get('JDBC_CONNECTION_URL')
        jdbc_conn_insert.jdbc_enforce_ssl = res_connection.get('ConnectionProperties').get('JDBC_ENFORCE_SSL')
        jdbc_conn_insert.master_username = res_connection.get('ConnectionProperties').get('USERNAME')
    if res_connection.get('PhysicalConnectionRequirements'):
        jdbc_conn_insert.network_availability_zone = res_connection.get('PhysicalConnectionRequirements').get('AvailabilityZone')
        jdbc_conn_insert.network_subnet_id = res_connection.get('PhysicalConnectionRequirements').get('SubnetId')
        jdbc_conn_insert.network_sg_id = "|".join(res_connection.get('PhysicalConnectionRequirements').get('SecurityGroupIdList'))
    jdbc_conn_insert.creation_time = res_connection.get('CreationTime')
    jdbc_conn_insert.last_updated_time = res_connection.get('LastUpdatedTime')
    res = crud.list_aws_jdbc_instance_source_by_account(jdbc_conn)
    if res:
        crud.update_jdbc_conn(jdbc_conn_insert)
    else:
        crud.add_jdbc_conn(jdbc_conn_insert)

def __glue(account: str, region: str):
    iam_role_name = crud.get_iam_role(account)
    assumed_role = sts.assume_role(
        RoleArn=f"{iam_role_name}",
        RoleSessionName="glue-connection"
    )
    credentials = assumed_role['Credentials']
    return boto3.client('glue',
                        aws_access_key_id=credentials['AccessKeyId'],
                        aws_secret_access_key=credentials['SecretAccessKey'],
                        aws_session_token=credentials['SessionToken'],
                        region_name=region
                        )

def __ec2(account: str, region: str):
    iam_role_name = crud.get_iam_role(account)
    assumed_role = sts.assume_role(
        RoleArn=f"{iam_role_name}",
        RoleSessionName="ec2-client"
    )
    credentials = assumed_role['Credentials']
    return boto3.client('ec2',
                        aws_access_key_id=credentials['AccessKeyId'],
                        aws_secret_access_key=credentials['SecretAccessKey'],
                        aws_session_token=credentials['SessionToken'],
                        region_name=region), credentials

def __lakeformation(account: str, region: str):
    iam_role_name = crud.get_iam_role(account)
    assumed_role = sts.assume_role(
        RoleArn=f"{iam_role_name}",
        RoleSessionName="lakeformation-client"
    )
    credentials = assumed_role['Credentials']
    return boto3.client('lakeformation',
                        aws_access_key_id=credentials['AccessKeyId'],
                        aws_secret_access_key=credentials['SecretAccessKey'],
                        aws_session_token=credentials['SessionToken'],
                        region_name=region)

def __create_jdbc_url(engine: str, host: str, port: str):
    # see https://docs.aws.amazon.com/glue/latest/dg/connection-properties.html#connection-properties-jdbc
    if engine == 'aurora-postgres' or engine == "postgres":
        return f"jdbc:postgresql://{host}:{port}/postgres"
    elif engine == 'aurora-mysql' or engine == "mysql":
        return f"jdbc:mysql://{host}:{port}/information_schema"
    elif engine == 'oracle':
        return f"jdbc:oracle:thin://@{host}:{port}/database"
    return ''

# Add S3 bucket, SQS queues access policies
def __update_access_policy_for_account():
    s3_resource = boto3.session.Session().resource('s3')
    # check if s3 bucket, sqs exists
    bucket_name = admin_bucket_name
    try:
        missing_resource = bucket_name
        s3_resource.meta.client.head_bucket(
            Bucket=bucket_name
        )
        sqs = boto3.client(
            'sqs',
            region_name=admin_region
        )
        missing_resource = f"{const.SOLUTION_NAME}-Crawler"
        sqs.get_queue_url(
            QueueName=f"{const.SOLUTION_NAME}-Crawler",
            QueueOwnerAWSAccountId=admin_account_id
        )
        missing_resource = f"{const.SOLUTION_NAME}-DiscoveryJob"
        sqs.get_queue_url(
            QueueName=f"{const.SOLUTION_NAME}-DiscoveryJob",
            QueueOwnerAWSAccountId=admin_account_id
        )
        missing_resource = f"{const.SOLUTION_NAME}-AutoSyncData"
        sqs.get_queue_url(
            QueueName=f"{const.SOLUTION_NAME}-AutoSyncData",
            QueueOwnerAWSAccountId=admin_account_id
        )
    except Exception as err:
        logger.error(f"Required resource {missing_resource} is missing, skipping region: {admin_region}")
        return

    accounts: list[Account] = crud.list_all_accounts_by_region(region=admin_region)
    if len(accounts) > 0:
        bucket_access_principals = []
        sqs_job_trigger_principals = []
        sqs_crawler_trigger_principals = []
        auto_sync_data_trigger_principals = []
        for account in accounts:

            role_arn = __gen_role_arn(account_id=account.account_id, region=account.region,
                                      role_name=_agent_role_name)
            # validate assumed
            if __assume_role(account.account_id, role_arn):
                s3_access_role_arn = f"arn:{partition}:iam::{account.account_id}:role/{const.SOLUTION_NAME}GlueDetectionJobRole-{account.region}"
                bucket_access_principals.append(s3_access_role_arn)

                crawler_trigger_role_arn = f"arn:{partition}:iam::{account.account_id}:role/{const.SOLUTION_NAME}RoleForCrawlerEvent-{account.region}"
                sqs_crawler_trigger_principals.append(crawler_trigger_role_arn)

                job_trigger_role_arn = f"arn:{partition}:iam::{account.account_id}:role/{const.SOLUTION_NAME}DiscoveryJobRole-{account.region}"
                sqs_job_trigger_principals.append(job_trigger_role_arn)

                auto_sync_data_role_arn = f"arn:{partition}:iam::{account.account_id}:role/{const.SOLUTION_NAME}DeleteAgentResourcesRole-{account.region}"
                auto_sync_data_trigger_principals.append(auto_sync_data_role_arn)

        # Bucket policies are limited to 20 KB in size.
        read_statement = {
            "Sid": "res-r",
            "Effect": "Allow",
            "Principal": {
                "AWS": bucket_access_principals
            },
            "Action": [
                "s3:GetObject",
            ],
            "Resource": [
                f"arn:{partition}:s3:::{bucket_name}/job/*",
                f"arn:{partition}:s3:::{bucket_name}/template/*"
            ]
        }
        list_statement = {
            "Sid": "res-l",
            "Effect": "Allow",
            "Principal": {
                "AWS": bucket_access_principals
            },
            "Action": [
                "s3:ListBucket",
            ],
            "Resource": [
                f"arn:{partition}:s3:::{bucket_name}"
            ]
        }
        write_statement = {
            "Sid": "res-w",
            "Effect": "Allow",
            "Principal": {
                "AWS": bucket_access_principals
            },
            "Action": [
                "s3:GetObject",
                "s3:PutObject",
            ],
            "Resource": f"arn:{partition}:s3:::{bucket_name}/glue-database/*"
        }
        s3 = boto3.client('s3')
        """ :type : pyboto3.s3 """
        try:
            restored_statements = []
            bucket_policy = json.loads(s3.get_bucket_policy(Bucket=bucket_name)['Policy'])
            for stat in bucket_policy['Statement']:
                if 'Sid' in stat and stat['Sid'] in ('res-w', 'res-r', 'res-l'):
                    continue
                else:
                    restored_statements.append(stat)

            restored_statements.append(read_statement)
            restored_statements.append(write_statement)
            restored_statements.append(list_statement)
            bucket_policy['Statement'] = restored_statements

            s3.put_bucket_policy(
                Bucket=bucket_name,
                Policy=json.dumps(bucket_policy)
            )
        except Exception as err:
            logger.error(traceback.format_exc())

        sqs = boto3.client(
            'sqs',
            region_name=admin_region
        )
        """ :type : pyboto3.sqs """
        sqs_job_trigger_policy = {
            "Version": "2008-10-17",
            "Statement": [
                {
                    "Effect": "Allow",
                    "Principal": {
                        "AWS": f"arn:{partition}:iam::{admin_account_id}:root"
                    },
                    "Action": "SQS:*",
                    "Resource": f"arn:{partition}:sqs:{admin_region}:{admin_account_id}:{const.SOLUTION_NAME}-DiscoveryJob"
                },
                {
                    "Effect": "Allow",
                    "Principal": {
                        "AWS": sqs_job_trigger_principals
                    },
                    "Action": "SQS:SendMessage",
                    "Resource": f"arn:{partition}:sqs:{admin_region}:{admin_account_id}:{const.SOLUTION_NAME}-DiscoveryJob"
                }
            ]
        }
        sqs_crawler_trigger_policy = {
            "Version": "2008-10-17",
            "Statement": [
                {
                    "Effect": "Allow",
                    "Principal": {
                        "AWS": f"arn:{partition}:iam::{admin_account_id}:root"
                    },
                    "Action": "SQS:*",
                    "Resource": f"arn:{partition}:sqs:{admin_region}:{admin_account_id}:{const.SOLUTION_NAME}-Crawler"
                },
                {
                    "Effect": "Allow",
                    "Principal": {
                        "AWS": sqs_crawler_trigger_principals
                    },
                    "Action": "SQS:SendMessage",
                    "Resource": f"arn:{partition}:sqs:{admin_region}:{admin_account_id}:{const.SOLUTION_NAME}-Crawler"
                }
            ]
        }
        auto_sync_data_trigger_policy = {
            "Version": "2008-10-17",
            "Statement": [
                {
                    "Effect": "Allow",
                    "Principal": {
                        "AWS": f"arn:{partition}:iam::{admin_account_id}:root"
                    },
                    "Action": "SQS:*",
                    "Resource": f"arn:{partition}:sqs:{admin_region}:{admin_account_id}:{const.SOLUTION_NAME}-AutoSyncData"
                },
                {
                    "Effect": "Allow",
                    "Principal": {
                        "AWS": auto_sync_data_trigger_principals
                    },
                    "Action": "SQS:SendMessage",
                    "Resource": f"arn:{partition}:sqs:{admin_region}:{admin_account_id}:{const.SOLUTION_NAME}-AutoSyncData"
                }
            ]
        }
        try:
            sqs.set_queue_attributes(
                QueueUrl=sqs.get_queue_url(QueueName=f"{const.SOLUTION_NAME}-DiscoveryJob")['QueueUrl'],
                Attributes={
                    "Policy": json.dumps(sqs_job_trigger_policy)
                }
            )
        except Exception as err:
            logger.error(traceback.format_exc())
        try:
            sqs.set_queue_attributes(
                QueueUrl=sqs.get_queue_url(QueueName=f"{const.SOLUTION_NAME}-Crawler")['QueueUrl'],
                Attributes={
                    "Policy": json.dumps(sqs_crawler_trigger_policy)
                }
            )
        except Exception as err:
            logger.error(traceback.format_exc())
        try:
            sqs.set_queue_attributes(
                QueueUrl=sqs.get_queue_url(QueueName=f"{const.SOLUTION_NAME}-AutoSyncData")["QueueUrl"],
                Attributes={
                    "Policy": json.dumps(auto_sync_data_trigger_policy)
                }
            )
        except Exception as err:
            logger.error(traceback.format_exc())


def __update_access_policy_for_ou(ou_id: str):
    pass


def __gen_role_arn(account_id: str, region: str, role_name: str):
    return f'arn:{partition}:iam::{account_id}:role/{const.SOLUTION_NAME}{role_name}-{region}'


def __assume_role(account_id: str, role_arn: str):
    try:
        response = sts.assume_role(
            RoleArn=role_arn,
            RoleSessionName='account_onboard_check',
            DurationSeconds=900,
        )
        return True
    except ClientError as e:
        if e.response['Error']['Code'] != 'AccessDenied':
            logger.info(e)
        return False


def __list_rds_schema(account, region, credentials, instance_name, payload, rds_security_groups, rds_subnet_id):
    logger.info("__list_rds_schema")
    function_name = f"{const.SOLUTION_NAME}-{instance_name[0:50]}"
    schema_path = []
    lambda_ = boto3.client(
        'lambda',
        aws_access_key_id=credentials['AccessKeyId'],
        aws_secret_access_key=credentials['SecretAccessKey'],
        aws_session_token=credentials['SessionToken'],
        region_name=region
    )
    logger.info("__list_rds_schema get lambda")
    """ :type : pyboto3.lambda_ """
    state = ''
    try:
        response = lambda_.get_function(FunctionName=function_name)
        # If lambda is not used for a long time, it will be in an inactive state and restored through invoke
        state = response['Configuration']['State']
        if state != 'Active':
            response = lambda_.invoke(
                FunctionName=function_name,
                InvocationType='Event',
                Payload="{}",
            )
    except Exception as e:
        logger.info("__list_rds_schema get lambda error and create")
        logger.info(str(e))
        lambda_.create_function(
            FunctionName=function_name,
            Handler='lambda_function.lambda_handler',
            Runtime='python3.9',
            Role=f"arn:{partition}:iam::{account}:role/SDPSLambdaRdsRole-{region}",
            Code={
                'S3Bucket': f"aws-gcr-solutions-{region}",
                'S3Key': 'aws-sensitive-data-protection/1.0.0/resource/schema_detection_function.zip',
            },
            Timeout=120,
            VpcConfig={
                'SubnetIds': [
                    rds_subnet_id,
                ],
                'SecurityGroupIds': rds_security_groups
            },
        )
    time_elapsed = 0
    timeout = 600
    while state != 'Active':
        sleep(SLEEP_TIME)
        time_elapsed += SLEEP_TIME
        if time_elapsed >= timeout:
            break
        response = lambda_.get_function(FunctionName=function_name)
        logger.info("__list_rds_schema get lambda function:")
        state = response['Configuration']['State']
        logger.info(state)

    logger.info("__list_rds_schema get lambda function invoke:")
    response = lambda_.invoke(
        FunctionName=function_name,
        InvocationType='RequestResponse',
        Payload=json.dumps(payload),
    )
    body = response['Payload'].read().decode()
    logger.info(body)
    # delete lambda
    logger.info("__list_rds_schema get lambda function delete:")
    logger.info(function_name)
    # lambda_.delete_function(FunctionName=function_name)
    schemas = json.loads(body)
    if schemas['statusCode'] == 500:
        raise Exception(schemas['errorMessage'])
    else:
        for schema in schemas['schema']:
            if schema['system'] == False:
                schema_path.append(schema['name'])
    logger.info("__list_rds_schema schema_path")
    logger.info(schema_path)
    return schema_path

def __delete_data_source_by_account(account_id: str, region: str):
    try:
        crud.delete_s3_bucket_source_by_account(account_id=account_id, region=region)
    except Exception:
        logger.error(traceback.format_exc())
    try:
        crud.delete_rds_instance_source_by_account(account_id=account_id, region=region)
    except Exception:
        logger.error(traceback.format_exc())

def __delete_account(account_id: str, region: str):
    try:
        crud.delete_account_by_region(account_id=account_id, region=region)
    except Exception:
        logger.error(traceback.format_exc())

def query_glue_connections(account: AccountInfo):
    res, list = [], []
    account_id, region = __get_admin_info(account)
    next_token = ""

    while True:
        response = __glue(account=account_id, region=region).get_connections(
            CatalogId=account_id,
            Filter={'ConnectionType': 'JDBC'},
            HidePassword=True,
            NextToken=next_token,
        )
        current_connections = response['ConnectionList']
        list.extend(current_connections)
        next_token = response.get('NextToken')

        if not next_token:
            break
    jdbc_list = query_jdbc_connections_sub_info()
    jdbc_dict = {item[0]: f"{convert_provider_id_2_name(item[1])}-{item[2]}" for item in jdbc_list}
    for item in list:
        if not item['Name'].startswith(const.SOLUTION_NAME):
            if item['Name'] in jdbc_dict:
                item['usedBy'] = jdbc_dict[item['Name']]
            res.append(item)
    return res

def query_jdbc_connections_sub_info():
    return crud.query_jdbc_connections_sub_info()

def list_buckets(account: AdminAccountInfo):
    _, region = __get_admin_info(account)
    iam_role_name = crud.get_iam_role(account.account_id)
    assumed_role = sts.assume_role(RoleArn=f"{iam_role_name}",
                                   RoleSessionName="glue-s3-connection")
    credentials = assumed_role['Credentials']
    s3 = boto3.client('s3',
                      aws_access_key_id=credentials['AccessKeyId'],
                      aws_secret_access_key=credentials['SecretAccessKey'],
                      aws_session_token=credentials['SessionToken'],
                      region_name=region)
    res = s3.list_buckets()
    return res["Buckets"]

def query_glue_databases(account: AdminAccountInfo):
    return __glue(account=account.account_id, region=account.region).get_databases()['DatabaseList']

def query_account_network(account: AccountInfo):
    account_id, region = __get_admin_info(account)
    ec2_client, __ = __ec2(account=account_id, region=region)
    vpcs = query_all_vpc(ec2_client)
    vpc_list = [{"vpcId": vpc.get('VpcId'), "name": gen_resource_name(vpc)} for vpc in vpcs]
    if account.account_provider_id != Provider.AWS_CLOUD.value:
        res = __query_third_account_network(vpc_list, ec2_client)
        return res
    else:
        return __query_aws_account_network(vpc_list, ec2_client)

# async def add_conn_jdbc_async(jdbcConn: JDBCInstanceSource):
#     key = f"{jdbcConn.account_provider_id}/{jdbcConn.account_id}/{jdbcConn.region}"
#     try:
#         add_jdbc_conn(jdbcConn)
#         return (key, "SUCCESSED", "")
#     except Exception as e:
#         return (key, "FAILED", str(e))

def __query_third_account_network(vpc_list, ec2_client: any):
    try:
        response = ec2_client.describe_security_groups(Filters=[
            {'Name': 'vpc-id', 'Values': [vpc["vpcId"] for vpc in vpc_list]},
            {'Name': 'group-name', 'Values': [const.SECURITY_GROUP_JDBC]}
        ])
        vpc_ids = [item['VpcId'] for item in response['SecurityGroups']]
        subnets = ec2_client.describe_subnets(Filters=[{'Name': 'vpc-id', 'Values': [vpc_ids[0]]}])['Subnets']
        selected_subnet = subnets
        subnets_str_from_env = os.getenv('SubnetIds', '')
        if subnets_str_from_env:
            subnets_from_env = subnets_str_from_env.split(',')
            selected_subnet = [item for item in subnets if item.get('SubnetId') in subnets_from_env]
        target_subnets = [{'subnetId': subnet["SubnetId"], 'arn': subnet["SubnetArn"], "subnetName": gen_resource_name(subnet)} for subnet in selected_subnet]
        vpc_info = ec2_client.describe_vpcs(VpcIds=[vpc_ids[0]])['Vpcs'][0]
        return {"vpcs": [{'vpcId': vpc_info['VpcId'],
                          'vpcName': [obj for obj in vpc_info['Tags'] if obj["Key"] == "Name"][0]["Value"],
                          'subnets': target_subnets,
                          'securityGroups': [{'securityGroupId': response['SecurityGroups'][0]['GroupId'],
                                              'securityGroupName': response['SecurityGroups'][0]['GroupName']}]}]
                }
    except ClientError as ce:
        logger.error(traceback.format_exc())
        if ce.response['Error']['Code'] == 'InvalidGroup.NotFound':
            raise BizException(MessageEnum.SOURCE_SECURITYGROUP_NOT_FOUND.get_code(),
                               MessageEnum.SOURCE_SECURITYGROUP_NOT_FOUND.get_msg())
        else:
            raise BizException(MessageEnum.BIZ_UNKNOWN_ERR.get_code(),
                               MessageEnum.BIZ_UNKNOWN_ERR.get_msg())
    except Exception as e:
        raise BizException(MessageEnum.BIZ_UNKNOWN_ERR.get_code(),
                           MessageEnum.BIZ_UNKNOWN_ERR.get_msg())

def __query_aws_account_network(vpc_list, ec2_client: any):
    res = []
    subnets = ec2_client.describe_subnets()['Subnets']
    securityGroups = ec2_client.describe_security_groups()['SecurityGroups']
    for vpc_info in vpc_list:
        vpc = {'vpcId': vpc_info["vpcId"], 'vpcName': vpc_info["name"]}
        vpc['subnets'] = [{"subnetId": subnet['SubnetId'],
                           "arn": subnet['SubnetArn'],
                           "subnetName": gen_resource_name(subnet)
                           } for subnet in subnets if subnet['VpcId'] == vpc_info["vpcId"]]
        vpc['securityGroups'] = [{"securityGroupId": security['GroupId'], "securityGroupName": security['GroupName']} for security in securityGroups if security['VpcId'] == vpc_info["vpcId"]]
        res.append(vpc)
    return {'vpcs': res}

def gen_resource_name(resource):
    if "Tags" in resource and "Name" in [tag["Key"] for tag in resource["Tags"]]:
        return [tag["Value"] for tag in resource["Tags"] if tag["Key"] == "Name"][0]
    else:
        return '-'

def test_glue_conn(account, connection):
    return boto3.client('glue').start_connection_test(
        CatalogId=account,
        ConnectionName=connection
    )['ConnectionTest']['Status']


def list_data_location():
    res = []
    provider_list = crud.list_distinct_provider()
    for item in provider_list:
        regions: list[SourceRegion] = crud.list_distinct_region_by_provider(item.id)
        accounts_db = crud.get_account_list_by_provider(item.id)
        if not regions:
            continue
        if not accounts_db:
            continue
        for subItem in regions:
            accounts = []
            for account in accounts_db:
                if account.region == subItem.region_name:
                    accounts.append(account)
            if len(accounts) == 0:
                continue
            location = DataLocationInfo()
            location.account_count = len(accounts)
            location.source = item.provider_name
            location.region = subItem.region_name
            location.coordinate = subItem.region_cord
            location.region_alias = subItem.region_alias
            location.provider_id = item.id
            res.append(location)
    res = sorted(res, key=lambda x: x.account_count, reverse=True)
    return res

def query_regions_by_provider(provider_id: int):
    return crud.query_regions_by_provider(provider_id)

def query_full_provider_resource_infos():
    res = []
    provider_list = crud.query_provider_list()
    for item in provider_list:
        provider = ProviderResourceFullInfo()
        provider.description = item.description
        provider.provider_id = item.id
        provider.provider_name = item.provider_name
        provider.resources = []
        resources = crud.query_resources_by_provider(item.id)
        for subItem in resources:
            resource = SourceResourceBase()
            resource.apply_region_ids = subItem.apply_region_ids
            resource.description = subItem.description
            resource.resource_alias = subItem.resource_alias
            resource.resource_name = subItem.resource_name
            resource.status = subItem.status
            provider.resources.append(resource)
        res.append(provider)
    return res

def list_providers():
    return crud.query_provider_list()

def get_db_names_4_jdbc(jdbc: JdbcSource, schemas: str):
    if not __validate_jdbc_url(jdbc.connection_url):
        raise BizException(MessageEnum.SOURCE_JDBC_URL_FORMAT_ERROR.get_code(),
                           MessageEnum.SOURCE_JDBC_URL_FORMAT_ERROR.get_msg())
    # list schemas
    db_names = set()
    if jdbc.connection_url.startswith('jdbc:mysql'):
        schemas = list_jdbc_databases(jdbc)
        return set(schemas)
    else:
        schema = get_schema_from_url(jdbc.connection_url)
        if schema:
            db_names.add(schema)
        if schemas:
            db_names.update(schemas.splitlines())
        if not db_names:
            raise BizException(MessageEnum.SOURCE_JDBC_JDBC_NO_DATABASE.get_code(),
                               MessageEnum.SOURCE_JDBC_JDBC_NO_DATABASE.get_msg())
        return db_names

def get_db_names(url: str, schemas: str):
    if not __validate_jdbc_url(url):
        raise BizException(MessageEnum.SOURCE_JDBC_URL_FORMAT_ERROR.get_code(),
                           MessageEnum.SOURCE_JDBC_URL_FORMAT_ERROR.get_msg())
    # list schemas
    db_names = set()
    schema = get_schema_from_url(url)
    if schema:
        db_names.add(schema)
    if schemas:
        db_names.update(schemas.splitlines())
    if not db_names:
        raise BizException(MessageEnum.SOURCE_JDBC_JDBC_NO_DATABASE.get_code(),
                           MessageEnum.SOURCE_JDBC_JDBC_NO_DATABASE.get_msg())
    return db_names

def get_schema_from_url(url):
    for pattern in _jdbc_url_patterns:
        match = re.match(pattern, url)
        if match:
            if match.groups():
                return match.groups()[0]
            return None


def grant_lake_formation_permission(credentials, crawler_role_arn, glue_database_name):
    lakeformation = boto3.client('lakeformation',
                                 aws_access_key_id=credentials['AccessKeyId'],
                                 aws_secret_access_key=credentials['SecretAccessKey'],
                                 aws_session_token=credentials['SessionToken'],
                                 region_name=admin_region)
    """ :type : pyboto3.lakeformation """
    # retry for grant permissions
    num_retries = GRANT_PERMISSIONS_RETRIES
    while num_retries > 0:
        try:
            response = lakeformation.grant_permissions(
                Principal={
                    'DataLakePrincipalIdentifier': f"{crawler_role_arn}"
                },
                Resource={
                    'Database': {
                        'Name': glue_database_name
                    }
                },
                Permissions=['ALL'],
                PermissionsWithGrantOption=['ALL']
            )
        except Exception as e:
            sleep(SLEEP_MIN_TIME)
            num_retries -= 1
        else:
            break
    else:
        raise BizException(MessageEnum.SOURCE_UNCONNECTED.get_code(), MessageEnum.SOURCE_UNCONNECTED.get_msg())


def query_connection_detail(account: JDBCInstanceSourceBase):
    account_id, region = __get_admin_info(account)
    source: JDBCInstanceSourceFullInfo = crud.get_jdbc_instance_source_glue(provider_id=account.account_provider_id,
                                                                            account=account.account_id,
                                                                            region=account.region,
                                                                            instance_id=account.instance_id)
    if not source or not source.glue_connection:
        raise BizException(MessageEnum.SOURCE_JDBC_CONNECTION_NOT_EXIST.get_code(),
                           MessageEnum.SOURCE_JDBC_CONNECTION_NOT_EXIST.get_msg())

    conn = __glue(account_id, region).get_connection(Name=source.glue_connection)['Connection']
    conn['ConnectionProperties']['JDBC_CONNECTION_SCHEMA'] = source.jdbc_connection_schema
    return conn

def __gen_resources_name(jdbc):
    database_type = convert_provider_id_2_database_type(jdbc.account_provider_id)
    glue_connection_name = f"{const.SOLUTION_NAME}-{database_type}-{jdbc.instance_id}"
    glue_database_name = f"{const.SOLUTION_NAME}-{database_type}-{jdbc.instance_id}"
    crawler_name = f"{const.SOLUTION_NAME}-{database_type}-{jdbc.instance_id}"
    return glue_connection_name, glue_database_name, crawler_name

def __get_excludes_file_exts():
    extensions = list(set([ext for extensions_list in const.UNSTRUCTURED_FILES.values() for ext in extensions_list]))
    return ["*.{" + ",".join(extensions) + "}"]

def __get_glue_client(account, region):
    iam_role_name = crud.get_iam_role(account)
    assumed_role = sts.assume_role(
        RoleArn=f"{iam_role_name}",
        RoleSessionName="glue-connection"
    )
    credentials = assumed_role['Credentials']
    glue = boto3.client('glue',
                        aws_access_key_id=credentials['AccessKeyId'],
                        aws_secret_access_key=credentials['SecretAccessKey'],
                        aws_session_token=credentials['SessionToken'],
                        region_name=region
                        )
    return glue

# def list_jdbc_databases(source: JdbcSource) -> list[str]:
#     url_arr = source.connection_url.split(":")
#     if len(url_arr) != 4:
#         raise BizException(MessageEnum.SOURCE_JDBC_URL_FORMAT_ERROR.get_code(), MessageEnum.SOURCE_JDBC_URL_FORMAT_ERROR.get_msg())
#     if url_arr[1] != "mysql":
#         raise BizException(MessageEnum.SOURCE_JDBC_LIST_DATABASES_NOT_SUPPORTED.get_code(), MessageEnum.SOURCE_JDBC_LIST_DATABASES_NOT_SUPPORTED.get_msg())
#     host = url_arr[2][2:]
#     port = int(url_arr[3].split("/")[0])
#     user = source.username
#     password = source.password
#     if source.secret_id:
#         secrets_client = boto3.client('secretsmanager')
#         secret_response = secrets_client.get_secret_value(SecretId=source.secret_id)
#         secrets = json.loads(secret_response['SecretString'])
#         user = secrets['username']
#         password = secrets['password']
#     mysql_database = jdbc_database.MySQLDatabase(host, port, user, password)
#     databases = mysql_database.list_databases()
#     logger.info(databases)
#     return databases

def batch_create(file: UploadFile = File(...)):
    time_str = time.time()
    jdbc_from_excel_set = set()
    created_jdbc_list = []
    account_set = set()
    # Check if the file is an Excel file
    if not file.filename.endswith('.xlsx'):
        raise BizException(MessageEnum.SOURCE_BATCH_CREATE_FORMAT_ERR.get_code(),
                           MessageEnum.SOURCE_BATCH_CREATE_FORMAT_ERR.get_msg())
    # Read the Excel file
    content = file.file.read()
    workbook = openpyxl.load_workbook(BytesIO(content), read_only=False)
    try:
        sheet = workbook.get_sheet_by_name(const.BATCH_SHEET)
    except KeyError:
        raise BizException(MessageEnum.SOURCE_BATCH_SHEET_NOT_FOUND.get_code(),
                           MessageEnum.SOURCE_BATCH_SHEET_NOT_FOUND.get_msg())
    header = [cell for cell in sheet.iter_rows(min_row=2, max_row=2, values_only=True)][0]
    sheet.delete_cols(12, amount=2)
    sheet.insert_cols(12, amount=2)
    sheet.cell(row=2, column=12, value="Result")
    sheet.cell(row=2, column=13, value="Details")
    accounts = crud.get_enable_account_list()
    accounts_list = [f"{account[0]}/{account[1]}/{account[2]}" for account in accounts]
    for row_index, row in enumerate(sheet.iter_rows(min_row=3), start=2):
        if all(cell.value is None for cell in row):
            continue
        res, msg = __check_empty_for_field(row, header)
        if res:
            __add_error_msg(sheet, row_index, msg)
        elif sheet.cell(row=row_index + 1, column=2).value not in [0, 1]:
            __add_error_msg(sheet, row_index, f"The value of {header[1]} must be 0 or 1")
        elif not __validate_jdbc_url(str(row[3].value)):
            __add_error_msg(sheet, row_index, f"The value of {header[3]} must be in the format jdbc:protocol://host:port")
        elif not str(row[3].value).startswith('jdbc:mysql') and not row[4].value:
            __add_error_msg(sheet, row_index, f"Non-MySQL-type data source {header[4]} cannot be null")
        elif len(str(row[2].value)) > const.CONNECTION_DESC_MAX_LEN:
            __add_error_msg(sheet, row_index, f"The value of {header[2]} must not exceed 2048")
        elif f"{row[10].value}/{row[8].value}/{row[9].value}/{row[0].value}" in jdbc_from_excel_set:
            __add_error_msg(sheet, row_index, f"The value of {header[0]}, {header[8]}, {header[9]}, {header[10]}  already exist in the preceding rows")
        elif f"{row[10].value}/{row[8].value}/{row[9].value}" not in accounts_list:
            __add_error_msg(sheet, row_index, "The account is not existed!")
        else:
            jdbc_from_excel_set.add(f"{row[10].value}/{row[8].value}/{row[9].value}/{row[0].value}")
            account_set.add(f"{row[10].value}/{row[8].value}/{row[9].value}")
            created_jdbc_list.append(__gen_created_jdbc(row))
    # Query network info
    if account_set:
        account_info = list(account_set)[0].split("/")
        network = query_account_network(AccountInfo(account_provider_id=account_info[0], account_id=account_info[1], region=account_info[2])) \
            .get('vpcs', [])[0]
        vpc_id = network.get('vpcId')
        subnets = [subnet.get('subnetId') for subnet in network.get('subnets')]
        security_group_id = network.get('securityGroups', [])[0].get('securityGroupId')
        created_jdbc_list = map_network_jdbc(created_jdbc_list, vpc_id, subnets, security_group_id)
        batch_result = asyncio.run(batch_add_conn_jdbc(created_jdbc_list))
        result = {f"{item[0]}/{item[1]}/{item[2]}/{item[3]}": f"{item[4]}/{item[5]}" for item in batch_result}
        for row_index, row in enumerate(sheet.iter_rows(min_row=3), start=2):
            if row[11].value:
                continue
            v = result.get(f"{row[10].value}/{row[8].value}/{row[9].value}/{row[0].value}")
            if v:
                if v.split('/')[0] == "SUCCESSED":
                    __add_success_msg(sheet, row_index)
                else:
                    __add_error_msg(sheet, row_index, v.split('/')[1])
    # Write into excel
    excel_bytes = BytesIO()
    workbook.save(excel_bytes)
    excel_bytes.seek(0)
    # Upload to S3
    batch_create_ds = f"{const.BATCH_CREATE_REPORT_PATH}/report_{time_str}.xlsx"
    __s3_client.upload_fileobj(excel_bytes, admin_bucket_name, batch_create_ds)
    return f'report_{time_str}'

def __check_empty_for_field(row, header):
    if row[0].value is None or str(row[0].value).strip() == const.EMPTY_STR:
        return True, f"{header[0]} should not be empty"
    if row[1].value is None or str(row[1].value).strip() == const.EMPTY_STR:
        return True, f"{header[1]} should not be empty"
    if row[3].value is None or str(row[3].value).strip() == const.EMPTY_STR:
        return True, f"{header[3]} should not be empty"
    if row[5].value is None or str(row[5].value).strip() == const.EMPTY_STR:
        if row[6].value is None or str(row[6].value).strip() == const.EMPTY_STR:
            return True, f"{header[6]} should not be empty when {header[5]} is empty"
        if row[7].value is None or str(row[7].value).strip() == const.EMPTY_STR:
            return True, f"{header[7]} should not be empty when {header[5]} is empty"
    if row[8].value is None or str(row[8].value).strip() == const.EMPTY_STR:
        return True, f"{header[8]} should not be empty"
    if row[9].value is None or str(row[9].value).strip() == const.EMPTY_STR:
        return True, f"{header[9]} should not be empty"
    if row[10].value is None or str(row[10].value).strip() == const.EMPTY_STR:
        return True, f"{header[10]} should not be empty"
    return False, None

def map_network_jdbc(created_jdbc_list: [JDBCInstanceSource], vpc_id, subnets, security_group_id):
    res = []
    for item in created_jdbc_list:
        item.network_sg_id = security_group_id
        item.network_subnet_id = random.choice(subnets)
        res.append(item)
    return res

def query_batch_status(filename: str):
    success, warning, failed = 0, 0, 0
    file_key = f"{const.BATCH_CREATE_REPORT_PATH}/{filename}.xlsx"
    response = __s3_client.list_objects_v2(Bucket=admin_bucket_name, Prefix=const.BATCH_CREATE_REPORT_PATH)
    for obj in response.get('Contents', []):
        if obj['Key'] == file_key:
            response = __s3_client.get_object(Bucket=admin_bucket_name, Key=file_key)
            excel_bytes = response['Body'].read()
            workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
            try:
                sheet = workbook[const.BATCH_SHEET]
            except KeyError:
                raise BizException(MessageEnum.SOURCE_BATCH_SHEET_NOT_FOUND.get_code(),
                                   MessageEnum.SOURCE_BATCH_SHEET_NOT_FOUND.get_msg())
            for _, row in enumerate(sheet.iter_rows(values_only=True, min_row=3)):
                if row[11] == "FAILED":
                    failed += 1
                if row[11] == "SUCCESSED":
                    success += 1
                if row[11] == "WARNING":
                    warning += 1
            return {"success": success, "warning": warning, "failed": failed}
    return 0

def download_batch_file(filename: str):
    key = f'{const.BATCH_CREATE_REPORT_PATH}/{filename}.xlsx'
    if filename.startswith("template-zh"):
        key = const.BATCH_CREATE_TEMPLATE_PATH_CN
    if filename.startswith("template-en"):
        key = const.BATCH_CREATE_TEMPLATE_PATH_EN
    url = __s3_client.generate_presigned_url(
        ClientMethod="get_object",
        Params={'Bucket': admin_bucket_name, 'Key': key},
        ExpiresIn=60
    )
    return url

def __add_error_msg(sheet, row_index, msg):
    if msg == const.EXISTED_MSG:
        sheet.cell(row=row_index + 1, column=12, value="WARNING")
        sheet.cell(row=row_index + 1, column=12).font = Font(color='563112', bold=True)
    else:
        sheet.cell(row=row_index + 1, column=12, value="FAILED")
        sheet.cell(row=row_index + 1, column=12).font = Font(color='FF0000', bold=True)
    sheet.cell(row=row_index + 1, column=13, value=msg)

def __add_success_msg(sheet, row_index):
    sheet.cell(row=row_index + 1, column=12, value="SUCCESSED")

def __gen_created_jdbc(row):
    created_jdbc = JDBCInstanceSource()
    created_jdbc.instance_id = row[0].value
    created_jdbc.jdbc_enforce_ssl = "true" if row[1].value == 1 else "false"
    created_jdbc.description = str(row[2].value)
    created_jdbc.jdbc_connection_url = str(row[3].value)
    created_jdbc.jdbc_connection_schema = str(row[4].value).replace(",", "\n") if row[4].value else const.EMPTY_STR
    created_jdbc.secret = str(row[5].value)
    created_jdbc.master_username = str(row[6].value)
    created_jdbc.password = str(row[7].value)
    created_jdbc.account_id = str(row[8].value)
    created_jdbc.region = str(row[9].value)
    created_jdbc.account_provider_id = row[10].value
    created_jdbc.creation_time = ""
    created_jdbc.custom_jdbc_cert = ""
    created_jdbc.custom_jdbc_cert_string = ""
    created_jdbc.jdbc_driver_class_name = ""
    created_jdbc.jdbc_driver_jar_uri = ""
    created_jdbc.last_updated_time = ""
    created_jdbc.network_availability_zone = ""
    created_jdbc.secret = ""
    created_jdbc.skip_custom_jdbc_cert_validation = "false"
    return created_jdbc

async def batch_add_conn_jdbc(created_jdbc_list):
    tasks = [asyncio.create_task(__add_jdbc_conn_batch(jdbc)) for jdbc in created_jdbc_list]
    return await asyncio.gather(*tasks)


def batch_sync_jdbc(jdbc_list):
    return asyncio.run(batch_sync_jdbc_manager(jdbc_list))

async def batch_sync_jdbc_manager(jdbc_list):
    tasks = [asyncio.create_task(__batch_sync_jdbc_worker(jdbc)) for jdbc in jdbc_list]
    return await asyncio.gather(*tasks)

async def __batch_sync_jdbc_worker(jdbc):
    sync_jdbc_connection(jdbc)
    # return jdbc.account_provider_id, jdbc.account_id, jdbc.region, jdbc.instance_id, "SUCCESSED", None
    # except BizException as be:
    #     return jdbc.account_provider_id, jdbc.account_id, jdbc.region, jdbc.instance_id, "FAILED", be.__msg__()
    # except Exception as e:
    #     return jdbc.account_provider_id, jdbc.account_id, jdbc.region, jdbc.instance_id, "FAILED", str(e)
# def get_schema_by_snapshot(provider_id: int, account_id: str, instance: str, region: str):
#     res = crud.get_schema_by_snapshot(provider_id, account_id, instance, region)
#     return res[0][0].replace(',', '\n').split('\n') if res else None, res[0][1] if res else None

# def get_schema_by_real_time(provider_id: int, account_id: str, instance: str, region: str, db_info: bool = False):
#     db, subnet_id = None, None
#     assume_account, assume_region = __get_admin_info(JDBCInstanceSourceBase(account_provider_id=provider_id, account_id=account_id, instance_id=instance, region=region))
#     connection_rds = crud.get_connection_by_instance(provider_id, account_id, instance, region)
#     glue = __get_glue_client(assume_account, assume_region)
#     connection = glue.get_connection(Name=connection_rds[0][0]).get('Connection', {})
#     if connection_rds[0] and connection_rds[0][0]:
#         subnet_id = connection.get('PhysicalConnectionRequirements', {}).get('SubnetId')
#     if db_info:
#         connection_properties = connection.get("ConnectionProperties", {})
#         jdbc_source = JdbcSource(username=connection_properties.get("USERNAME"),
#                                  password=connection_properties.get("PASSWORD"),
#                                  secret_id=connection_properties.get("SECRET_ID"),
#                                  connection_url=connection_properties.get("JDBC_CONNECTION_URL")
#                                  )
#         db = list_jdbc_databases(jdbc_source)
#     return db, subnet_id

# def sync_schema_by_job(provider_id: int, account_id: str, instance: str, region: str, schema: str):
#     jdbc = JDBCInstanceSourceBase(instance_id=instance, account_provider_id=provider_id, account_id=account_id, region=region)
#     account_id, region = __get_admin_info(jdbc)
#     # Query Info
#     info = crud.get_crawler_glue_db_by_instance(provider_id, account_id, instance, region)
#     if not info:
#         return
#     crawler_role_arn = __gen_role_arn(account_id=account_id,
#                                       region=region,
#                                       role_name='GlueDetectionJobRole')
#     db_names = schema.split("\n")
#     jdbc_targets = __gen_jdbc_targets_from_db_names(info[0][2], db_names)
#     # Update Crawler
#     __update_crawler(provider_id, account_id, instance, region, jdbc_targets, info[0][0], info[0][1], crawler_role_arn)
#     # Update RDS
#     crud.update_schema_by_account(provider_id, account_id, instance, region, schema)

def __gen_jdbc_targets_from_db_names(connection_name, db_names):
    jdbc_targets = []
    for db_name in db_names:
        trimmed_db_name = db_name.strip()
        if trimmed_db_name:
            jdbc_targets.append({
                'ConnectionName': connection_name,
                'Path': f"{trimmed_db_name}/%"
            })
    return jdbc_targets

def __update_crawler(provider_id, account_id, instance, region, jdbc_targets, crawler_name, glue_database, crawler_role_arn):
    assume_account, assume_region = __get_admin_info(JDBCInstanceSourceBase(account_provider_id=provider_id,
                                                                            account_id=account_id,
                                                                            instance_id=instance,
                                                                            region=region))
    try:
        __get_glue_client(assume_account, assume_region).update_crawler(
            Name=crawler_name,
            Role=crawler_role_arn,
            DatabaseName=glue_database,
            Targets={
                'JdbcTargets': jdbc_targets,
            },
            SchemaChangePolicy={
                'UpdateBehavior': 'UPDATE_IN_DATABASE',
                'DeleteBehavior': 'DELETE_FROM_DATABASE'
            }
        )
    except Exception as e:
        logger.error(traceback.format_exc())
        raise BizException(MessageEnum.BIZ_UNKNOWN_ERR.get_code(),
                           MessageEnum.BIZ_UNKNOWN_ERR.get_msg())

def __get_admin_info(jdbc):
    account_id = jdbc.account_id if jdbc.account_provider_id == Provider.AWS_CLOUD.value else admin_account_id
    region = jdbc.region if jdbc.account_provider_id == Provider.AWS_CLOUD.value else admin_region
    return account_id, region

async def __add_jdbc_conn_batch(jdbc: JDBCInstanceSource):
    try:
        add_jdbc_conn(jdbc)
        return jdbc.account_provider_id, jdbc.account_id, jdbc.region, jdbc.instance_id, "SUCCESSED", None
    except BizException as be:
        return jdbc.account_provider_id, jdbc.account_id, jdbc.region, jdbc.instance_id, "FAILED", be.__msg__()
    except Exception as e:
        return jdbc.account_provider_id, jdbc.account_id, jdbc.region, jdbc.instance_id, "FAILED", str(e)
