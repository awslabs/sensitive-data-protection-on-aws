import asyncio
from io import BytesIO
import json
import os
import tempfile
import time
import boto3
from fastapi import File, UploadFile
import openpyxl
from common.constant import const
from common.response_wrapper import S3WrapEncoder
from common.exception_handler import BizException
from common.enum import MessageEnum, DatabaseType, IdentifierDependency, IdentifierType, OperationType
from common.query_condition import QueryCondition
from common.reference_parameter import admin_bucket_name
from catalog.service_dashboard import get_database_by_identifier
from common import concurrent_upload2s3
from common.abilities import insert_error_msg_2_cells, insert_success_2_cells
from template import schemas, crud


caller_identity = boto3.client('sts').get_caller_identity()
__s3_client = boto3.client('s3')

def get_identifiers(condition: QueryCondition):
    return crud.get_identifiers(condition)


def get_identifiers_by_template(tid: int):
    res = []
    tmp = crud.get_identifiers_by_template(tid)
    for item in tmp:
        res.append(item['identifier_id'])
    return res


def get_identifier(id: int):
    return crud.get_identifier(id)

def create_identifier(identifier: schemas.TemplateIdentifier):
    res_list = crud.get_identify_by_name(identifier.name)
    if res_list:
        raise BizException(MessageEnum.TEMPLATE_IDENTIFIER_EXISTS.get_code(),
                           MessageEnum.TEMPLATE_IDENTIFIER_EXISTS.get_msg())
    check_rule(identifier)
    return crud.create_identifier(identifier)


def delete_identifier(id: int):
    ref = []
    identifier = crud.get_identifier(id)
    ref_templates = crud.get_mappings_by_identifier(id)
    ref_buckets = get_database_by_identifier(identifier.name, DatabaseType.S3.value)
    ref_rds = get_database_by_identifier(identifier.name, DatabaseType.RDS.value)
    if ref_templates:
        ref.append(IdentifierDependency.TEMPLATE.value)
    if ref_buckets:
        ref.append(f'{IdentifierDependency.S3.value}:{len(ref_buckets)}')
    if ref_rds:
        ref.append(f'{IdentifierDependency.RDS.value}:{len(ref_rds)}')
    if not ref:
        crud.delete_identifier(id)
    else:
        raise BizException(MessageEnum.TEMPLATE_IDENTIFIER_USED.get_code(), MessageEnum.TEMPLATE_IDENTIFIER_USED.get_msg(), ref)


def update_identifier(id: int, identifier: schemas.TemplateIdentifier):
    res_list = crud.get_identify_by_name(identifier.name)
    if res_list and res_list[0].id != id:
        raise BizException(MessageEnum.TEMPLATE_IDENTIFIER_EXISTS.get_code(),
                           MessageEnum.TEMPLATE_IDENTIFIER_EXISTS.get_msg())
    check_rule(identifier)
    snapshot_no, res = crud.update_identifier(id, identifier)
    # used_by_template = crud.get_mappings_by_identifier(id)
    if snapshot_no:
        sync_s3(snapshot_no)
    return res


def get_template(id: int):
    # MVP：single template, and the template with id 1 is returned by default
    return crud.get_template(const.DEFAULT_TEMPLATE_ID)


def create_mapping(mapping: schemas.TemplateMapping):
    snapshot_no = crud.create_mapping(mapping)
    sync_s3(snapshot_no)
    return snapshot_no


def update_mapping(id: int, mapping: schemas.TemplateMapping):
    snapshot_no, res = crud.update_mapping(id, mapping)
    sync_s3(snapshot_no)
    return res


def delete_mapping(ids: list[int]):
    snapshot_no = crud.delete_mapping(ids)
    sync_s3(snapshot_no)


def get_mappings(condition: QueryCondition):
    # MVP：single template, and the template with id 1 is returned by default
    return crud.get_mappings(const.DEFAULT_TEMPLATE_ID, condition)


def get_template_snapshot_no(id: int):
    return crud.get_template_snapshot_no(id)


def get_props_by_type(tid: int):
    return crud.get_props_by_type(tid)


def create_prop(prop: schemas.TemplateIdentifierProp):
    res_list = crud.get_props_by_name_and_type(prop)
    if res_list:
        raise BizException(MessageEnum.TEMPLATE_PROPS_EXISTS.get_code(), MessageEnum.TEMPLATE_PROPS_EXISTS.get_msg())
    return crud.create_pop(prop)


def delete_prop(id: int):
    refs = crud.get_refs_by_prop(id)
    if refs:
        raise BizException(MessageEnum.TEMPLATE_PROPS_USED.get_code(), MessageEnum.TEMPLATE_PROPS_USED.get_msg())
    crud.delete_prop(id)


def update_prop(id: int, prop: schemas.TemplateIdentifierProp):
    res_list = crud.get_props_by_name_and_type(prop)
    if res_list:
        raise BizException(MessageEnum.TEMPLATE_PROPS_EXISTS.get_code(), MessageEnum.TEMPLATE_PROPS_EXISTS.get_msg())
    snapshot_no, res = crud.update_prop(id, prop)
    if snapshot_no:
        sync_s3(snapshot_no)
    return res


def sync_s3(snapshot_no):
    res_json = {}
    identifiers = []
    # generate new version
    template = crud.get_template(const.DEFAULT_TEMPLATE_ID)
    res = crud.get_ref_identifiers(const.DEFAULT_TEMPLATE_ID)
    for item in res:
        item_json = {}
        item_json['name'] = item[4]
        item_json['classification'] = item[5]
        item_json['rule'] = item[6]
        item_json['type'] = item[7]
        item_json['privacy'] = item[8]
        item_json['header_keywords'] = json.loads(item[9]) if item[9] else item[9]
        if item[11]:
            item_json['exclude_keywords'] = [val for val in json.loads(item[11]) if val]
        else:
            item_json['exclude_keywords'] = []
        item_json['description'] = item[10]
        item_json['max_distance'] = item[12]
        item_json['min_occurrence'] = item[13]
        identifiers.append(item_json)
    json_str = S3WrapEncoder.convert(template, ['id', 'name'])
    res_json['template_id'] = json_str['id']
    res_json['template_name'] = json_str['name']
    res_json['identifiers'] = identifiers
    # upload
    client = boto3.client('s3')
    client.put_object(
        Body=json.dumps(res_json, ensure_ascii=False),
        Bucket=admin_bucket_name,
        Key='template/template-{}-{}.json'.format(res_json['template_id'], snapshot_no),
    )


def check_rule(identifier: schemas.TemplateIdentifier):
    if identifier.type == IdentifierType.CUSTOM.value and identifier.rule == const.EMPTY_STR:
        raise BizException(MessageEnum.TEMPLATE_IDENTIFIER_RULES_EMPTY.get_code(),
                           MessageEnum.TEMPLATE_IDENTIFIER_RULES_EMPTY.get_msg())
    if identifier.type == IdentifierType.CUSTOM.value and identifier.header_keywords and '""' in identifier.header_keywords[1:-1].split(","):
        raise BizException(MessageEnum.TEMPLATE_HEADER_KEYWORDS_EMPTY.get_code(),
                           MessageEnum.TEMPLATE_HEADER_KEYWORDS_EMPTY.get_msg())


def export_identify(key):
    default_sheet = "Sheet"
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("IDENTIFY", index=0)
    if default_sheet in workbook.sheetnames and len(workbook.sheetnames) > 1:
        origin_sheet = workbook[default_sheet]
        workbook.remove(origin_sheet)
    sheet.append(const.EXPORT_IDENTIFY_HEADER)
    result = crud.get_all_identifiers()
    for row_num, row_data in enumerate(result, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            if col_num == 3:
                cell_value = "Customize" if cell_value == 1 else "Built in"
            if col_num == 4:
                cell_value = "Machine learning" if cell_value == 0 else "Regex"
            if col_num == 5:
                cell_value = "Non PII" if cell_value == 0 else "PII"
            sheet.cell(row=row_num, column=col_num).value = cell_value
    workbook.active = 0
    file_name = f"identify_{key}.xlsx"
    tmp_file = f"{tempfile.gettempdir()}/{file_name}"
    report_file = f"{const.IDENTIFY_REPORT}/{file_name}"
    workbook.save(tmp_file)
    stats = os.stat(tmp_file)
    if stats.st_size < 6 * 1024 * 1024:
        __s3_client.upload_file(tmp_file, admin_bucket_name, report_file)
    else:
        concurrent_upload2s3(admin_bucket_name, report_file, tmp_file, __s3_client)
    os.remove(tmp_file)
    method_parameters = {'Bucket': admin_bucket_name, 'Key': report_file}
    pre_url = __s3_client.generate_presigned_url(
        ClientMethod="get_object",
        Params=method_parameters,
        ExpiresIn=60
    )
    return pre_url


def delete_report(key):
    __s3_client.delete_object(Bucket=admin_bucket_name, Key=f"{const.BATCH_CREATE_IDENTIFIER_REPORT_PATH}/{key}.xlsx")


def query_batch_status(filename: str):
    success, warning, failed = 0, 0, 0
    file_key = f"{const.BATCH_CREATE_IDENTIFIER_REPORT_PATH}/{filename}.xlsx"
    response = __s3_client.list_objects_v2(Bucket=admin_bucket_name, Prefix=const.BATCH_CREATE_IDENTIFIER_REPORT_PATH)
    for obj in response.get('Contents', []):
        if obj['Key'] == file_key:
            response = __s3_client.get_object(Bucket=admin_bucket_name, Key=file_key)
            excel_bytes = response['Body'].read()
            workbook = openpyxl.load_workbook(BytesIO(excel_bytes))
            try:
                sheet = workbook[const.BATCH_SHEET]
            except KeyError:
                raise BizException(MessageEnum.SOURCE_BATCH_SHEET_NOT_FOUND.get_code(),
                                   MessageEnum.SOURCE_BATCH_SHEET_NOT_FOUND.get_msg())
            for _, row in enumerate(sheet.iter_rows(values_only=True, min_row=3)):
                if row[9] == "FAILED":
                    failed += 1
                if row[9] == "SUCCESSED":
                    success += 1
                if row[9] == "WARNING":
                    warning += 1
            return {"success": success, "warning": warning, "failed": failed}
    return 0

def download_batch_file(filename: str):
    key = f'{const.BATCH_CREATE_IDENTIFIER_REPORT_PATH}/{filename}.xlsx'
    if filename.startswith("identifier-template-zh"):
        key = const.BATCH_CREATE_IDENTIFIER_TEMPLATE_PATH_CN
    if filename.startswith("identifier-template-en"):
        key = const.BATCH_CREATE_IDENTIFIER_TEMPLATE_PATH_EN
    url = __s3_client.generate_presigned_url(
        ClientMethod="get_object",
        Params={'Bucket': admin_bucket_name, 'Key': key},
        ExpiresIn=60
    )
    return url

def batch_create(file: UploadFile = File(...)):
    res_column_index = 10
    time_str = time.time()
    identifier_from_excel_set = set()
    created_identifier_list = []
    category_list = crud.get_props_by_type(1)
    label_list = crud.get_props_by_type(2)
    # Check if the file is an Excel file
    if not file.filename.endswith('.xlsx'):
        raise BizException(MessageEnum.SOURCE_BATCH_CREATE_FORMAT_ERR.get_code(),
                           MessageEnum.SOURCE_BATCH_CREATE_FORMAT_ERR.get_msg())
    # Read the Excel file
    content = file.file.read()
    workbook = openpyxl.load_workbook(BytesIO(content), read_only=False)
    try:
        sheet = workbook.get_sheet_by_name(const.BATCH_SHEET)
    except KeyError:
        raise BizException(MessageEnum.SOURCE_BATCH_SHEET_NOT_FOUND.get_code(),
                           MessageEnum.SOURCE_BATCH_SHEET_NOT_FOUND.get_msg())
    # if sheet.max_row == 2:
    #     raise BizException(MessageEnum.SOURCE_BATCH_SHEET_NO_CONTENT.get_code(),
    #                        MessageEnum.SOURCE_BATCH_SHEET_NO_CONTENT.get_msg())
    header = [cell for cell in sheet.iter_rows(min_row=2, max_row=2, values_only=True)][0]
    sheet.delete_cols(10, amount=2)
    sheet.insert_cols(10, amount=2)
    sheet.cell(row=2, column=10, value="Result")
    sheet.cell(row=2, column=11, value="Details")
    identifiers = crud.get_all_identifiers()
    identifier_list = [identifier[0] for identifier in identifiers]
    no_content = True
    for row_index, row in enumerate(sheet.iter_rows(min_row=3), start=2):
        props = []
        if all(cell.value is None for cell in row):
            continue
        no_content = False
        res, msg = __check_empty_for_field(row, header)
        if res:
            insert_error_msg_2_cells(sheet, row_index, msg, res_column_index)
        elif f"{row[0].value}" in identifier_from_excel_set:
            insert_error_msg_2_cells(sheet, row_index, f"The value of {header[0]} already exist in the preceding rows", res_column_index)
        elif not row[3].value and not row[4].value:
            # Content validation rules and title keywords validation rules cannot be empty at the same time.
            insert_error_msg_2_cells(sheet, row_index, f"The value of {header[3]} and {header[4]} cannot be empty at the same time.", res_column_index)   
        elif not __is_pos_int_or_none(row[5]):
            insert_error_msg_2_cells(sheet, row_index, f"The value of {header[5]} must be positive integer.", res_column_index) 
        elif not __is_pos_int_or_none(row[6]):
            insert_error_msg_2_cells(sheet, row_index, f"The value of {header[6]} must be positive integer.", res_column_index) 
        elif row[7].value and not [category for category in category_list if category.prop_name.lower() == row[7].value.strip().lower()]:
            # category = 
            # if category:
            #     props.append(category[0])
            # else:
            insert_error_msg_2_cells(sheet, row_index, f"The value of {header[7]} is not existed in System, please take a check", res_column_index)
        elif row[8].value and not [label for label in label_list if label.prop_name.lower() == row[8].value.strip().lower()]:
            # label = 
            # if label:
            #     props.append(label[0])
            # else:
            insert_error_msg_2_cells(sheet, row_index, f"The value of {header[8]} is not existed in System, please take a check", res_column_index)
        elif row[0].value in identifier_list:
            # Account.account_provider_id, Account.account_id, Account.region
            insert_error_msg_2_cells(sheet, row_index, "A data identifier with the same name already exists", res_column_index)
        else:
            identifier_from_excel_set.add(row[0].value)
            if row[7].value:
                categories = [category for category in category_list if category.prop_name.lower() == row[7].value.strip().lower()]
                props.append(categories[0].id)
            if row[8].value:
                labels = [label for label in label_list if label.prop_name.lower() == row[8].value.strip().lower()]
                props.append(labels[0].id)
            # account_set.add(f"{row[10].value}/{row[8].value}/{row[9].value}")
            created_identifier_list.append(__gen_created_identifier(row, props))
    if no_content:
        raise BizException(MessageEnum.SOURCE_BATCH_SHEET_NO_CONTENT.get_code(),
                           MessageEnum.SOURCE_BATCH_SHEET_NO_CONTENT.get_msg())
    batch_result = asyncio.run(batch_add_identifier(created_identifier_list))
    result = {item[0]: f"{item[1]}/{item[2]}" for item in batch_result}
    for row_index, row in enumerate(sheet.iter_rows(min_row=3), start=2):
        # print(f"row[10] id {row[1].value} ")
        if row[10] and row[10].value:
            continue
        v = result.get(row[0].value)
        if v:
            if v.split('/')[0] == "SUCCESSED":
                insert_success_2_cells(sheet, row_index, res_column_index)
            else:
                insert_error_msg_2_cells(sheet, row_index, v.split('/')[1], res_column_index)
    # Write into excel
    excel_bytes = BytesIO()
    workbook.save(excel_bytes)
    excel_bytes.seek(0)
    # Upload to S3
    batch_create_ds = f"{const.BATCH_CREATE_IDENTIFIER_REPORT_PATH}/report_{time_str}.xlsx"
    __s3_client.upload_fileobj(excel_bytes, admin_bucket_name, batch_create_ds)
    return f'report_{time_str}'

def __check_empty_for_field(row, header):
    if row[0].value is None or str(row[0].value).strip() == const.EMPTY_STR:
        return True, f"{header[0]} should not be empty"
    if row[2].value is None or str(row[2].value).strip() == const.EMPTY_STR:
        return True, f"{header[2]} should not be empty"
    return False, None

def __gen_created_identifier(row, props):
    created_identifier = schemas.TemplateIdentifier()
    created_identifier.name = row[0].value
    created_identifier.description = str(row[1].value)
    created_identifier.props = props
    created_identifier.rule = row[2].value
    created_identifier.header_keywords = json.dumps(str(row[3].value).split(",")) if row[3].value else None
    created_identifier.exclude_keywords = json.dumps(str(row[4].value).split(",")) if row[4].value else None
    created_identifier.max_distance = row[5].value
    created_identifier.min_occurrence = row[6].value
    created_identifier.type = IdentifierType.CUSTOM.value
    return created_identifier

def __is_pos_int_or_none(cell):
    if not cell or not cell.value:
        return True
    try:
        if int(cell.value) > 0:
            return True
    except Exception as e:
        return False
    return False

async def batch_add_identifier(created_identifier_list):
    tasks = [asyncio.create_task(__add_create_identifier_batch(identifier)) for identifier in created_identifier_list]
    return await asyncio.gather(*tasks)

async def __add_create_identifier_batch(identifier: schemas.TemplateIdentifier):
    try:
        create_identifier(identifier)
        return identifier.name, "SUCCESSED", None
    except BizException as be:
        return identifier.name, "FAILED", be.__msg__()
    except Exception as e:
        return identifier.name, "FAILED", str(e)
