from common.enum import (Provider,
                         ProviderName,
                         DatabaseType)
from common.reference_parameter import logger, admin_account_id
from common.constant import const
from openpyxl.styles import Font, PatternFill


def convert_database_type_2_provider(database_type: str) -> int:
    if database_type == DatabaseType.JDBC_GOOGLE.value:
        return Provider.GOOGLE_CLOUD.value
    elif database_type == DatabaseType.JDBC_TENCENT.value:
        return Provider.TENCENT_CLOUD.value
    elif database_type == DatabaseType.JDBC_PROXY.value:
        return Provider.JDBC_PROXY.value
    else:
        return Provider.AWS_CLOUD.value


def convert_provider_id_2_database_type(provider: int) -> str:
    if provider == Provider.TENCENT_CLOUD.value:
        return DatabaseType.JDBC_TENCENT.value
    elif provider == Provider.GOOGLE_CLOUD.value:
        return DatabaseType.JDBC_GOOGLE.value
    elif provider == Provider.JDBC_PROXY.value:
        return DatabaseType.JDBC_PROXY.value
    else:
        return DatabaseType.JDBC_AWS.value


def convert_provider_id_2_name(provider: int) -> str:
    if provider == Provider.TENCENT_CLOUD.value:
        return ProviderName.TENCENT_CLOUD.value
    elif provider == Provider.GOOGLE_CLOUD.value:
        return ProviderName.GOOGLE_CLOUD.value
    elif provider == Provider.JDBC_PROXY.value:
        return ProviderName.JDBC_PROXY.value
    else:
        return ProviderName.AWS_CLOUD.value


# This function is used for detecting the glue connection/crawler is created for third-party JDBC connections,
# it only returns true when the JDBC connection is a third-party JDBC connection.
def need_change_account_id(database_type: str) -> bool:
    if database_type.startswith(DatabaseType.JDBC.value) and database_type != DatabaseType.JDBC_AWS.value:
        return True
    return False


def is_run_in_admin_vpc(database_type: str, account_id: str = None) -> bool:
    if database_type == DatabaseType.JDBC_AWS.value:
        return account_id == admin_account_id
    elif database_type.startswith(DatabaseType.JDBC.value):
        return True
    return False


def query_all_vpc(ec2_client):
    vpcs = []
    response = ec2_client.describe_vpcs()
    vpcs.append(response['Vpcs'])
    while 'NextToken' in response:
        response = ec2_client.describe_vpcs(NextToken=response['NextToken'])
        vpcs.append(response['Vpcs'])
    return vpcs[0]

def insert_error_msg_2_cells(sheet, row_index, msg, res_column_index):
    if msg == const.EXISTED_MSG:
        sheet.cell(row=row_index + 1, column=res_column_index, value="WARNING")
        sheet.cell(row=row_index + 1, column=res_column_index).font = Font(color='563112', bold=True)
    else:
        sheet.cell(row=row_index + 1, column=res_column_index, value="FAILED")
        sheet.cell(row=row_index + 1, column=res_column_index).font = Font(color='FF0000', bold=True)
    sheet.cell(row=row_index + 1, column=res_column_index + 1, value=msg)

def insert_success_2_cells(sheet, row_index, res_column_index):
    sheet.cell(row=row_index + 1, column=res_column_index, value="SUCCESSED")