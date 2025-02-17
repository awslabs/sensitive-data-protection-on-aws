import os
import logging
import boto3
import json
import db.models_discovery_job as models
from . import crud, schemas
from common.exception_handler import BizException
from common.enum import MessageEnum, JobState, RunState, RunDatabaseState, DatabaseType, AthenaQueryState
from common.constant import const
from common.query_condition import QueryCondition
from common.reference_parameter import logger, admin_account_id, admin_region, admin_bucket_name, partition, url_suffix, public_account_id, admin_subnet_ids
import traceback
import tools.mytime as mytime
import datetime, time, pytz
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from catalog.service import sync_job_detection_result
from common.abilities import need_change_account_id, convert_database_type_2_provider, is_run_in_admin_vpc
import config.service as config_service
from data_source import jdbc_schema
from tools import list_tool
from data_source.resource_list import list_resources_by_database_type

version = os.getenv(const.VERSION, '')
controller_function_name = os.getenv("ControllerFunctionName", f"{const.SOLUTION_NAME}-Controller")
sqs_resource = boto3.resource('sqs')

sql_result = "SELECT database_type,account_id,region,database_name,location,column_name,identifiers,sample_data FROM job_detection_output_table where run_id='%d' and privacy = 1"
sql_error = "SELECT account_id,region,database_type,database_name,table_name,error_message FROM job_detection_error_table where run_id='%d'"
extra_py_files = f"s3://{admin_bucket_name}/job/script/job_extra_files.zip"
report_key_template = "report/report-%d-%d.xlsx"


def list_jobs(condition: QueryCondition):
    return crud.list_jobs(condition)


def get_job(id: int):
    return crud.get_job(id)


def __deduplicate_list_of_objects(lst):
    deduplicated_set = set(lst)
    deduplicated_list = list(deduplicated_set)
    return deduplicated_list


def create_job(job: schemas.DiscoveryJobCreate):
    job.databases = __deduplicate_list_of_objects(job.databases)
    if job.depth_structured is None:
        job.depth_structured = 0
    if job.depth_unstructured is None:
        job.depth_unstructured = 0
    db_job = crud.create_job(job)
    if db_job.schedule != const.ON_DEMAND:
        create_event(db_job.id, db_job.schedule)
    return db_job


def create_event(job_id: int, schedule: str):
    rule_name = f'{const.SOLUTION_NAME}-Controller-{job_id}'
    client_events = boto3.client('events')
    response = client_events.put_rule(
        Name=rule_name,
        ScheduleExpression=schedule,
        State='ENABLED',
        Description=f'create by {const.SOLUTION_NAME}',
        Tags=[
            {
                'Key': const.TAG_KEY,
                'Value': const.TAG_VALUE
            },
            {
                'Key': 'JobId',
                'Value': str(job_id)
            },
        ],
    )

    input = {const.CONTROLLER_ACTION: const.CONTROLLER_ACTION_SCHEDULE_JOB, "JobId": job_id}
    response = client_events.put_targets(
        Rule=rule_name,
        Targets=[
            {
                'Id': '1',
                'Arn': f'arn:{partition}:lambda:{admin_region}:{admin_account_id}:function:{controller_function_name}',
                'Input': json.dumps(input),
            },
        ]
    )

    client_lambda = boto3.client('lambda')
    response = client_lambda.add_permission(
        Action='lambda:InvokeFunction',
        FunctionName=controller_function_name,
        Principal='events.amazonaws.com',
        SourceArn=f'arn:{partition}:events:{admin_region}:{admin_account_id}:rule/{rule_name}',
        StatementId=rule_name,
    )


def delete_job(id: int):
    db_job = crud.get_job(id)
    if db_job.state == JobState.RUNNING.value or db_job.state == JobState.OD_RUNNING.value:
        raise BizException(MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_JOB.get_code(),
                           MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_JOB.get_msg())
    delete_event(id)
    crud.delete_job(id)


def delete_event(job_id: int):
    rule_name = f'{const.SOLUTION_NAME}-Controller-{job_id}'
    client_events = boto3.client('events')
    try:
        response = client_events.remove_targets(
            Rule=rule_name,
            Ids=['1'],
            Force=True
        )
    except client_events.exceptions.ResourceNotFoundException as e:
        logger.warning(e)

    response = client_events.delete_rule(
        Name=rule_name,
    )

    try:
        client_lambda = boto3.client('lambda')
        response = client_lambda.remove_permission(
            FunctionName=controller_function_name,
            StatementId=rule_name,
        )
    except client_lambda.exceptions.ResourceNotFoundException as e:
        logger.warning(e)


def update_event(job_id: int, schedule: str):
    rule_name = f'{const.SOLUTION_NAME}-Controller-{job_id}'
    client_events = boto3.client('events')
    response = client_events.put_rule(
        Name=rule_name,
        ScheduleExpression=schedule,
        Description=f'create by {const.SOLUTION_NAME}',
    )


def update_job(id: int, job: schemas.DiscoveryJobUpdate):
    length = len(job.dict(exclude_unset=True))
    if length == 0:
        raise BizException(MessageEnum.DISCOVERY_JOB_INVALID_PARAMETER.get_code(),
                           MessageEnum.DISCOVERY_JOB_INVALID_PARAMETER.get_msg())
    if job.schedule:
        db_job = crud.get_job(id)
        if db_job.schedule == const.ON_DEMAND and job.schedule == const.ON_DEMAND:
            pass
        elif db_job.schedule == const.ON_DEMAND and job.schedule != const.ON_DEMAND:
            create_event(id, job.schedule)
        elif db_job.schedule != const.ON_DEMAND and job.schedule == const.ON_DEMAND:
            delete_event(id)
        elif db_job.schedule != job.schedule:
            update_event(id, job.schedule)
    crud.update_job(id, job)


def last_job_time() -> str:
    return crud.last_job_time()


def enable_job(id: int):
    db_job = crud.get_job(id)
    if db_job.schedule == const.ON_DEMAND:
        raise BizException(MessageEnum.DISCOVERY_JOB_CAN_CHANGE_STATE.get_code(),
                           MessageEnum.DISCOVERY_JOB_CAN_CHANGE_STATE.get_msg())
    if db_job.state != JobState.PAUSED.value:
        raise BizException(MessageEnum.DISCOVERY_JOB_CAN_ENABLE.get_code(),
                           MessageEnum.DISCOVERY_JOB_CAN_ENABLE.get_msg())
    rule_name = f'{const.SOLUTION_NAME}-Controller-{id}'
    client_events = boto3.client('events')
    client_events.enable_rule(Name=rule_name)
    crud.enable_job(id)


def disable_job(id: int):
    db_job = crud.get_job(id)
    if db_job.schedule == const.ON_DEMAND:
        raise BizException(MessageEnum.DISCOVERY_JOB_CAN_CHANGE_STATE.get_code(),
                           MessageEnum.DISCOVERY_JOB_CAN_CHANGE_STATE.get_msg())
    if db_job.state != JobState.IDLE.value:
        raise BizException(MessageEnum.DISCOVERY_JOB_CAN_DISABLE.get_code(),
                           MessageEnum.DISCOVERY_JOB_CAN_DISABLE.get_msg())
    rule_name = f'{const.SOLUTION_NAME}-Controller-{id}'
    client_events = boto3.client('events')
    client_events.disable_rule(Name=rule_name)
    crud.disable_job(id)


def start_job(job_id: int):
    run_id = crud.init_run(job_id)
    if run_id >= 0:
        run = crud.get_run(run_id)
        if not run.databases:
            crud.complete_run(run_id)
            raise BizException(MessageEnum.DISCOVERY_JOB_DATABASE_IS_EMPTY.get_code(),
                               MessageEnum.DISCOVERY_JOB_DATABASE_IS_EMPTY.get_msg())
        failed_run_database_count = __start_run_databases(run.databases)
        if failed_run_database_count == len(run.databases):
            crud.complete_run(run_id)
            raise BizException(MessageEnum.DISCOVERY_JOB_ALL_RUN_FAILED.get_code(),
                               MessageEnum.DISCOVERY_JOB_ALL_RUN_FAILED.get_msg())


def start_sample_job(job_id: int, table_name: str):
    run_id = crud.init_run(job_id)
    logger.info(run_id)
    if run_id >= 0:
        __start_sample_run(job_id, run_id, table_name)


def __get_job_number(database_type: str) -> int:
    if database_type in [DatabaseType.S3.value, DatabaseType.GLUE.value]:
        return int(config_service.get_config(const.CONFIG_SUB_JOB_NUMBER_S3, const.CONFIG_SUB_JOB_NUMBER_S3_DEFAULT_VALUE))
    return int(config_service.get_config(const.CONFIG_SUB_JOB_NUMBER_RDS, const.CONFIG_SUB_JOB_NUMBER_RDS_DEFAULT_VALUE))


def get_run_database_ip_count(database_type: str) -> int:
    crawler_ip = 0
    if database_type.startswith(DatabaseType.JDBC.value):
        crawler_ip = 3
    return crawler_ip + __get_job_number(database_type) * 2  # Each GlueJob requires 2 IPs


def __count_run_database_by_subnet() -> dict:
    count_run_database = {}
    run_databases = crud.get_running_run_databases()
    for run_database in run_databases:
        if not need_change_account_id(run_database.database_type):
            continue
        provider_id = convert_database_type_2_provider(run_database.database_type)
        _, subnet_id = jdbc_schema.get_schema_by_real_time(provider_id, run_database.account_id, run_database.region, run_database.database_name)
        count = count_run_database.get(subnet_id, 0)
        count_run_database[subnet_id] = count + 1
    logger.info(f"count_run_database:{count_run_database}")
    return count_run_database


def __enable_event_bridge(rule_name: str):
    client_events = boto3.client('events')
    client_events.enable_rule(Name=rule_name)


def __start_run_databases(run_databases):
    job_dic = {}
    run_dic = {}
    module_path = f's3://{admin_bucket_name}/job/ml-asset/python-module/'
    wheels = ["humanfriendly-10.0-py2.py3-none-any.whl",
              "protobuf-4.22.1-cp37-abi3-manylinux2014_x86_64.whl",
              "flatbuffers-23.3.3-py2.py3-none-any.whl",
              "coloredlogs-15.0.1-py2.py3-none-any.whl",
              "onnxruntime-1.13.1-cp310-cp310-manylinux_2_17_x86_64.manylinux2014_x86_64.whl",
              "sdpsner-1.0.0-py3-none-any.whl",
              ]
    limit_concurrency = False

    account_loop_wait = {}
    for run_database in run_databases:
        if is_run_in_admin_vpc(run_database.database_type, run_database.account_id):
            limit_concurrency = True
        account_id = run_database.account_id
        if need_change_account_id(run_database.database_type):
            account_id = admin_account_id
        if account_id in account_loop_wait:
            tmp = account_loop_wait[account_id]
            tmp = tmp + const.JOB_INTERVAL_WAIT
            account_loop_wait[account_id] = tmp
        else:
            account_loop_wait[account_id] = const.JOB_INTERVAL_WAIT

    if limit_concurrency:
        concurrent_run_job_number = int(config_service.get_config(const.CONFIG_CONCURRENT_RUN_JOB_NUMBER, const.CONFIG_CONCURRENT_RUN_JOB_NUMBER_DEFAULT_VALUE))
        logger.debug(f"concurrent_run_job_number:{concurrent_run_job_number}")
        count_run_database = __count_run_database_by_subnet()
        for key in account_loop_wait:
            if account_loop_wait[key] > const.JOB_INTERVAL_WAIT * concurrent_run_job_number:
                account_loop_wait[key] = const.JOB_INTERVAL_WAIT * concurrent_run_job_number

    job_placeholder = ","
    account_first_wait = {}
    failed_run_database_count = 0
    check_pending_started = False
    for run_database in run_databases:
        try:
            if is_run_in_admin_vpc(run_database.database_type, run_database.account_id):
                logger.debug(f"database_name:{run_database.database_name}")
                provider_id = convert_database_type_2_provider(run_database.database_type)
                database_schemas_real_time, subnet_id = jdbc_schema.get_schema_by_real_time(provider_id, run_database.account_id, run_database.region, run_database.database_name, True)
                count = count_run_database.get(subnet_id, 0)
                logger.debug(f"subnet_id:{subnet_id}")
                logger.debug(f"count:{count}")
                if count >= concurrent_run_job_number:
                    run_database.state = RunDatabaseState.PENDING.value
                    logger.debug(f"{run_database.database_name} break")
                    if not check_pending_started:
                        check_pending_started = True
                        __enable_event_bridge(f"{const.SOLUTION_NAME}-CheckPending")
                    continue
                logger.debug(f"run_database.database_name add")
                count_run_database[subnet_id] = count + 1
                if database_schemas_real_time:
                    database_schemas_snapshot, _ = jdbc_schema.get_schema_by_snapshot(provider_id, run_database.account_id, run_database.region, run_database.database_name)
                    logger.info(f'database_schemas_real_time:{database_schemas_real_time}')
                    logger.info(f'database_schemas_snapshot:{database_schemas_snapshot}')
                    if not list_tool.compare(database_schemas_real_time, database_schemas_snapshot):
                        jdbc_schema.sync_schema_by_job(provider_id, run_database.account_id, run_database.region, run_database.database_name, database_schemas_real_time)
                        logger.info(f'Updated schema:{database_schemas_real_time}')
                else:
                    logger.info(f'Unable to obtain the schema for {run_database.database_name}')

            run = run_dic.get(run_database.run_id)
            if not run:
                run = crud.get_run(run_database.run_id)
                run_dic[run_database.run_id] = run
            job = job_dic.get(run.job_id)
            if not job:
                job = crud.get_job(run.job_id)
                job_dic[run.job_id] = job

            account_id = run_database.account_id
            region = run_database.region
            if need_change_account_id(run_database.database_type):
                account_id = admin_account_id
                region = admin_region
            if account_id in account_first_wait:
                tmp = account_first_wait[account_id]
                tmp = tmp + const.JOB_INTERVAL_WAIT
                account_first_wait[account_id] = tmp
            else:
                account_first_wait[account_id] = 0
            # job_bookmark_option = "job-bookmark-enable" if job.range == 1 else "job-bookmark-disable"
            base_time = str(datetime.datetime.min)
            if job.range == 1 and run_database.base_time:
                base_time = mytime.format_time(run_database.base_time)
            need_run_crawler = True
            if run_database.database_type == DatabaseType.GLUE.value or run_database.table_name:
                need_run_crawler = False
            crawler_name = f"{const.SOLUTION_NAME}-{run_database.database_type}-{run_database.database_name}"
            glue_database_name = f"{const.SOLUTION_NAME}-{run_database.database_type}-{run_database.database_name}"
            if run_database.database_type == DatabaseType.GLUE.value:
                glue_database_name = run_database.database_name
            job_name_structured = f"{const.SOLUTION_NAME}-{run_database.database_type}-{run_database.database_name}"
            job_name_unstructured = f"{const.SOLUTION_NAME}-{DatabaseType.S3_UNSTRUCTURED.value}-{run_database.database_name}"
            run_name = f'{const.SOLUTION_NAME}-{run_database.run_id}-{run_database.id}-{run_database.uuid}'
            # agent_bucket_name = f"{const.AGENT_BUCKET_NAME_PREFIX}-{run_database.account_id}-{run_database.region}"
            unstructured_parser_job_image_uri = f"{public_account_id}.dkr.ecr.{run_database.region}.amazonaws.com{url_suffix}/aws-sensitive-data-protection-models:v1.1.2"
            unstructured_parser_job_role = f"arn:{partition}:iam::{run_database.account_id}:role/{const.SOLUTION_NAME}UnstructuredParserRole-{run_database.region}"
            execution_input = {
                "RunName": run_name,
                "JobNameStructured": job_name_structured,
                "JobNameUnstructured": job_name_unstructured,
                "NeedRunCrawler": need_run_crawler,
                "CrawlerName": crawler_name,
                "JobId": str(job.id),  # When calling Glue Job using StepFunction, the parameter must be of string type
                "RunId": str(run_database.run_id),
                "RunDatabaseId": str(run_database.id),
                "AccountId": run_database.account_id,  # The original account id is required here
                "Region": run_database.region,  # The original region is required here
                "DatabaseType": run_database.database_type,
                "DatabaseName": run_database.database_name,
                "GlueDatabaseName": glue_database_name,
                "UnstructuredDatabaseName": f"{const.SOLUTION_NAME}-{DatabaseType.S3_UNSTRUCTURED.value}-{run_database.database_name}",
                "TableName": run_database.table_name if run_database.table_name else job_placeholder,
                "TemplateId": str(run.template_id),
                "TemplateSnapshotNo": str(run.template_snapshot_no),
                "DepthStructured": "0" if run.depth_structured is None else str(run.depth_structured),
                "DepthUnstructured": "0" if run.depth_unstructured is None else str(run.depth_unstructured),
                "ExcludeKeywords": run.exclude_keywords if run.exclude_keywords else job_placeholder,
                "IncludeKeywords": run.include_keywords if run.include_keywords else job_placeholder,
                "ExcludeFileExtensions": run.exclude_file_extensions if run.exclude_file_extensions else job_placeholder,
                "IncludeFileExtensions": run.include_file_extensions if run.include_file_extensions else job_placeholder,
                "BaseTime": base_time,
                # "JobBookmarkOption": job_bookmark_option,
                "DetectionThreshold": str(job.detection_threshold),
                "OverWrite": str(job.overwrite),
                # "AgentBucketName": agent_bucket_name,
                "AdminAccountId": admin_account_id,
                "AdminBucketName": admin_bucket_name,
                "AdditionalPythonModules": ','.join([module_path + w for w in wheels]),
                "ExtraPyFiles": extra_py_files,
                "FirstWait": str(account_first_wait[account_id]),
                "LoopWait": str(account_loop_wait[account_id]),
                "JobNumber": __get_job_number(run_database.database_type),
                "QueueUrl": f'https://sqs.{region}.amazonaws.com{url_suffix}/{admin_account_id}/{const.SOLUTION_NAME}-DiscoveryJob',
                "UnstructuredParserJobImageUri": unstructured_parser_job_image_uri,
                "UnstructuredParserJobRole": unstructured_parser_job_role,
            }
            run_database.start_time = mytime.get_time()
            __create_job(run_database.database_type, account_id, region, run_database.database_name, job_name_structured, 'glue-job.py')
            if run_database.database_type == DatabaseType.S3.value:
                __create_job(run_database.database_type, account_id, region, run_database.database_name, job_name_unstructured, 'glue-job-unstructured.py')
            __exec_run(execution_input)
            run_database.state = RunDatabaseState.RUNNING.value
        except Exception:
            failed_run_database_count += 1
            msg = traceback.format_exc()
            run_database.state = RunDatabaseState.FAILED.value
            run_database.end_time = mytime.get_time()
            run_database.error_log = msg
            logger.exception("Run StepFunction exception:%s" % msg)
    crud.save_run_databases(run_databases)
    return failed_run_database_count


def __start_sample_run(job_id: int, run_id: int, table_name: str):
    job = crud.get_job(job_id)
    run = crud.get_run(run_id)
    run_databases = run.databases
    logger.info(run_databases)
    for run_database in run_databases:
        try:
            # job_bookmark_option = "job-bookmark-enable" if job.range == 1 else "job-bookmark-disable"
            base_time = str(datetime.datetime.min)
            if job.range == 1 and run_database.base_time is not None:
                base_time = mytime.format_time(run_database.base_time)
            job_name = f"{const.SOLUTION_NAME}-Sample-Job-S3"
            if run_database.database_type == DatabaseType.RDS.value:
                job_name = f"{const.SOLUTION_NAME}-Sample-Job-RDS" + run_database.database_name
            execution_input = {
                "--JobName": job_name,
                "--JobId": str(job.id),
                "--RunId": str(run_id),
                "--Limit": str(const.SAMPLE_LIMIT),
                "--AccountId": run_database.account_id,
                "--Region": run_database.region,
                "--DatabaseType": run_database.database_type,
                "--DatabaseName": run_database.database_name,
                "--TableName": table_name,
                "--Depth": str(job.depth_structured),
                "--BaseTime": base_time,
                "--BucketName": admin_bucket_name,
            }
            run_database.start_time = mytime.get_time()
            __create_job(run_database.database_type, run_database.account_id, run_database.region, run_database.database_name, job_name, 'glue-sample-job.py')
            __exec_sample_run(execution_input)
            run_database.state = RunDatabaseState.RUNNING.value
        except Exception as e:
            msg = traceback.format_exc()
            run_database.state = RunDatabaseState.FAILED.value
            run_database.end_time = mytime.get_time()
            run_database.error_log = msg
            logger.info(str(e))
            logger.exception("start_sample_run exception:%s" % msg)
    crud.save_run_databases(run_databases)


def __create_job(database_type: str, account_id, region, database_name, job_name, script_name):
    client_sts = boto3.client('sts')
    assumed_role_object = client_sts.assume_role(
        RoleArn=f'arn:{partition}:iam::{account_id}:role/{const.SOLUTION_NAME}RoleForAdmin-{region}',
        RoleSessionName="AssumeRoleSession2"
    )
    credentials = assumed_role_object['Credentials']

    client_glue = boto3.client('glue',
                               aws_access_key_id=credentials['AccessKeyId'],
                               aws_secret_access_key=credentials['SecretAccessKey'],
                               aws_session_token=credentials['SessionToken'],
                               region_name=region,
                               )
    try:
        client_glue.get_job(JobName=job_name)
    except client_glue.exceptions.EntityNotFoundException as e:
        if database_type == DatabaseType.RDS.value or database_type.startswith(DatabaseType.JDBC.value):
            # The imported connection name is not standardized
            response = client_glue.get_crawler(Name=f'{const.SOLUTION_NAME}-{database_type}-{database_name}')
            jdbc_targets = response["Crawler"]["Targets"]["JdbcTargets"]
            connection_set = set()
            if jdbc_targets:
                for jdbc_target in jdbc_targets:
                    connection_set.add(jdbc_target["ConnectionName"])
            client_glue.create_job(Name=job_name,
                                   Role=f'{const.SOLUTION_NAME}GlueDetectionJobRole-{region}',
                                   GlueVersion='4.0',
                                   Command={'Name': 'glueetl',
                                            'ScriptLocation': f's3://{admin_bucket_name}/job/script/{script_name}'},
                                   Tags={const.TAG_KEY: const.TAG_VALUE,
                                         const.TAG_ADMIN_ACCOUNT_ID: admin_account_id,
                                         const.VERSION: version},
                                   NumberOfWorkers=2,
                                   WorkerType='G.1X',
                                   ExecutionProperty={'MaxConcurrentRuns': 100},
                                   Timeout=30 * 24 * 60,
                                   Connections={'Connections': list(connection_set)},
                                   )
        else:
            client_glue.create_job(Name=job_name,
                                   Role=f'{const.SOLUTION_NAME}GlueDetectionJobRole-{region}',
                                   GlueVersion='4.0',
                                   Command={'Name': 'glueetl',
                                            'ScriptLocation': f's3://{admin_bucket_name}/job/script/{script_name}'},
                                   Tags={const.TAG_KEY: const.TAG_VALUE,
                                         const.TAG_ADMIN_ACCOUNT_ID: admin_account_id,
                                         const.VERSION: version},
                                   NumberOfWorkers=2,
                                   WorkerType='G.1X',
                                   ExecutionProperty={'MaxConcurrentRuns': 1000},
                                   Timeout=30 * 24 * 60,
                                   )


def __check_sfn_version(client_sfn, arn, account_id):
    response = client_sfn.list_tags_for_resource(resourceArn=arn)
    agent_version = 'Version Placeholder'
    for tag in response['tags']:
        if tag.get('key') == const.VERSION:
            agent_version = tag.get('value')
            break
    logger.info(f"{account_id} version is:{agent_version}")
    # Only check if the solution version is consistent.
    # Do not determine if the build version is consistent
    if os.getenv(const.MODE) != const.MODE_DEV:
        agent_solution_version = agent_version.split('-')[0]
        if not version.startswith(agent_solution_version):
            raise BizException(MessageEnum.DISCOVERY_JOB_AGENT_MISMATCHING_VERSION.get_code(),
                               MessageEnum.DISCOVERY_JOB_AGENT_MISMATCHING_VERSION.get_msg())


def __exec_run(execution_input):
    account_id = execution_input["AccountId"]
    region = execution_input["Region"]
    if need_change_account_id(execution_input["DatabaseType"]):
        account_id = execution_input["AdminAccountId"]
        region = admin_region
    client_sts = boto3.client('sts')
    assumed_role_object = client_sts.assume_role(
        RoleArn=f'arn:{partition}:iam::{account_id}:role/{const.SOLUTION_NAME}RoleForAdmin-{region}',
        RoleSessionName="AssumeRoleSession1"
    )
    credentials = assumed_role_object['Credentials']

    client_sfn = boto3.client('stepfunctions',
                              aws_access_key_id=credentials['AccessKeyId'],
                              aws_secret_access_key=credentials['SecretAccessKey'],
                              aws_session_token=credentials['SessionToken'],
                              region_name=region,
                              )
    arn = f'arn:{partition}:states:{region}:{account_id}:stateMachine:{const.SOLUTION_NAME}-DiscoveryJob'
    __check_sfn_version(client_sfn, arn, account_id)
    client_sfn.start_execution(
        stateMachineArn=arn,
        name=execution_input["RunName"],
        input=json.dumps(execution_input),
    )


def __exec_sample_run(execution_input):
    client_sts = boto3.client('sts')
    assumed_role_object = client_sts.assume_role(
        RoleArn=f'arn:{partition}:iam::{execution_input["--AccountId"]}:role/{const.SOLUTION_NAME}RoleForAdmin-{execution_input["--Region"]}',
        RoleSessionName="AssumeRoleSession1"
    )
    credentials = assumed_role_object['Credentials']
    client_glue = boto3.client('glue',
                               aws_access_key_id=credentials['AccessKeyId'],
                               aws_secret_access_key=credentials['SecretAccessKey'],
                               aws_session_token=credentials['SessionToken'],
                               region_name=execution_input["--Region"],
                               )
    logger.info(execution_input)
    response = client_glue.start_job_run(JobName=execution_input["--JobName"],
                                         Arguments=execution_input)
    logger.info(response)


def stop_job(job_id: int):
    db_run: models.DiscoveryJobRun = crud.get_running_run(job_id)
    if db_run is None:
        raise BizException(MessageEnum.DISCOVERY_JOB_NON_RUNNING.get_code(),
                           MessageEnum.DISCOVERY_JOB_NON_RUNNING.get_msg())
    if db_run.state == RunState.STOPPING.value:
        delta_seconds = (mytime.get_now() - db_run.modify_time).seconds
        if delta_seconds < const.LAMBDA_MAX_RUNTIME:
            raise BizException(MessageEnum.DISCOVERY_JOB_STOPPING.get_code(),
                               MessageEnum.DISCOVERY_JOB_STOPPING.get_msg())

    run_databases: list[models.DiscoveryJobRunDatabase] = db_run.databases
    crud.stop_run(job_id, db_run.id)
    job = crud.get_job(job_id)
    for run_database in run_databases:
        logger.info(f"Stop job,JobId:{job_id},RunId:{run_database.run_id},RunDatabaseId:{run_database.id},"
                    f"AccountId:{run_database.account_id},Region:{run_database.region},"
                    f"DatabaseType:{run_database.database_type},DatabaseName:{run_database.database_name}")
        __stop_step_function(run_database)
        __send_complete_run_database_message(job_id, run_database.run_id, run_database.id, run_database.account_id,
                                             run_database.region, run_database.database_type, run_database.database_name,
                                             job.overwrite == 1, RunDatabaseState.STOPPED.value)


def __stop_step_function(run_database: models.DiscoveryJobRunDatabase):
    account_id = run_database.account_id
    region = run_database.region
    if need_change_account_id(run_database.database_type):
        account_id = admin_account_id
        region = admin_region
    client_sts = boto3.client('sts')
    assumed_role_object = client_sts.assume_role(
        RoleArn=f'arn:{partition}:iam::{account_id}:role/{const.SOLUTION_NAME}RoleForAdmin-{region}',
        RoleSessionName="AssumeRoleSession1"
    )
    credentials = assumed_role_object['Credentials']

    client_sfn = boto3.client('stepfunctions',
                              aws_access_key_id=credentials['AccessKeyId'],
                              aws_secret_access_key=credentials['SecretAccessKey'],
                              aws_session_token=credentials['SessionToken'],
                              region_name=region,
                              )
    try:
        client_sfn.stop_execution(
            executionArn=f'arn:{partition}:states:{region}:{account_id}:execution:{const.SOLUTION_NAME}-DiscoveryJob:{const.SOLUTION_NAME}-{run_database.run_id}-{run_database.id}-{run_database.uuid}',
            )
    except client_sfn.exceptions.ExecutionDoesNotExist as e:
        logger.warning(e)


def get_runs(job_id: int):
    return crud.get_runs(job_id)


def get_run(run_id: int):
    return crud.get_run(run_id)


def list_run_databases_pagination(run_id: int, condition: QueryCondition):
    return crud.list_run_databases_pagination(run_id, condition)


def get_run_status(job_id: int, run_id: int) -> schemas.DiscoveryJobRunDatabaseStatus:
    run_list = crud.list_run_databases(run_id)
    success_count = fail_count = ready_count = pending_count = running_count = stopped_count = not_existed_count = 0
    success_per = fail_per = ready_per = pending_per = running_per = stopped_per = not_existed_per = 0

    total_count = len(run_list)
    if total_count > 0:
        for run_item in run_list:
            if run_item.state == RunDatabaseState.SUCCEEDED.value:
                success_count += 1
            elif run_item.state == RunDatabaseState.FAILED.value:
                fail_count += 1
            elif run_item.state == RunDatabaseState.READY.value:
                ready_count += 1
            elif run_item.state == RunDatabaseState.PENDING.value:
                pending_count += 1
            elif run_item.state == RunDatabaseState.RUNNING.value:
                running_count += 1
            elif run_item.state == RunDatabaseState.STOPPED.value:
                stopped_count += 1
            elif run_item.state == RunDatabaseState.NOT_EXIST.value:
                not_existed_count += 1

        fail_per = int(fail_count / total_count * 100)
        ready_per = int(ready_count / total_count * 100)
        pending_per = int(pending_count / total_count * 100)
        running_per = int(running_count / total_count * 100)
        stopped_per = int(stopped_count / total_count * 100)
        not_existed_per = int(not_existed_count / total_count * 100)
        success_per = 100 - fail_per - ready_per - running_per - stopped_per - not_existed_per
    status = schemas.DiscoveryJobRunDatabaseStatus(
        id=job_id,
        run_id=run_id,
        success_count=success_count,
        fail_count=fail_count,
        ready_count=ready_count,
        pending_count=pending_count,
        running_count=running_count,
        stopped_count=stopped_count,
        not_existed_count=not_existed_count,
        total_count=total_count,
        success_per=success_per,
        fail_per=fail_per,
        ready_per=ready_per,
        pending_per=pending_per,
        running_per=running_per,
        stopped_per=stopped_per,
        not_existed_per=not_existed_per
    )
    return status


def get_run_progress(job_id: int, run_id: int) -> list[schemas.DiscoveryJobRunDatabaseProgress]:
    job = crud.get_job(job_id)
    run = crud.get_run(run_id)
    run_current_table_count = __get_run_current_table_count(run_id)
    run_progress = []
    for run_database in run.databases:
        try:
            if run_database.state == RunDatabaseState.PENDING.value:
                progress = schemas.DiscoveryJobRunDatabaseProgress(run_database_id=run_database.id,
                                                                   current_table_count=-1,
                                                                   table_count=-1,
                                                                   current_table_count_unstructured=-1,
                                                                   table_count_unstructured=-1)
                run_progress.append(progress)
                continue
            base_time = datetime.datetime.min
            if job.range == 1 and run_database.base_time:
                base_time = run_database.base_time
            base_time = pytz.timezone('UTC').localize(base_time)

            current_table_count = run_current_table_count.get(run_database.id)
            if current_table_count is None:
                current_table_count = 0
            if run_database.table_name:
                table_count = len(run_database.table_name.split(","))
            else:
                table_count = __get_table_count_from_agent(run_database, base_time)
            if current_table_count > table_count:
                table_count = current_table_count

            table_count_unstructured = -1
            current_table_count_unstructured = -1
            if run_database.database_type == DatabaseType.S3.value and run.depth_unstructured != 0:
                current_table_count_unstructured = run_current_table_count.get(f"{run_database.id}-{DatabaseType.S3_UNSTRUCTURED.value}")
                if current_table_count_unstructured is None:
                    current_table_count_unstructured = 0
                table_count_unstructured = __get_table_count_from_agent(run_database, base_time, False)
            if current_table_count_unstructured > table_count_unstructured:
                table_count_unstructured = current_table_count_unstructured

            progress = schemas.DiscoveryJobRunDatabaseProgress(run_database_id=run_database.id,
                                                               current_table_count=current_table_count,
                                                               table_count=table_count,
                                                               current_table_count_unstructured=current_table_count_unstructured,
                                                               table_count_unstructured=table_count_unstructured)
            run_progress.append(progress)
        except Exception:
            message = traceback.format_exc()
            logger.exception(f"get table count from agent exception:{message}")
            progress = schemas.DiscoveryJobRunDatabaseProgress(run_database_id=run_database.id,
                                                               current_table_count=-1,
                                                               table_count=-1,
                                                               current_table_count_unstructured=-1,
                                                               table_count_unstructured=-1)
            run_progress.append(progress)
    return run_progress


def __get_run_current_table_count(run_id: int):
    sql = f"select run_database_id,database_type,count(distinct table_name) from sdps_database.job_detection_output_table where run_id='{run_id}' group by run_database_id,database_type"
    current_table_count = __query_athena(sql)
    if logger.isEnabledFor(logging.DEBUG):
        logger.debug(current_table_count)
    table_count = {}
    for row in current_table_count[1:]:
        row_result = [__get_cell_value(cell) for cell in row]
        key = int(row_result[0]) if row_result[1] != DatabaseType.S3_UNSTRUCTURED.value else f"{row_result[0]}-{DatabaseType.S3_UNSTRUCTURED.value}"
        table_count[key] = int(row_result[2])
    if logger.isEnabledFor(logging.DEBUG):
        logger.debug(table_count)
    return table_count


def __get_current_table_count(run_database_id: int):
    sql = f"select count(distinct table_name) from sdps_database.job_detection_output_table where run_database_id='{run_database_id}'"
    current_table_count = __query_athena(sql)
    logger.debug(current_table_count)
    return int(current_table_count[1][0]["VarCharValue"])


def __get_table_count_from_agent(run_database: models.DiscoveryJobRunDatabase, base_time: datetime.datetime, is_structured=True):
    client_sts = boto3.client('sts')
    account_id = run_database.account_id
    region = run_database.region
    if need_change_account_id(run_database.database_type):
        account_id = admin_account_id
        region = admin_region
    assumed_role_object = client_sts.assume_role(
        RoleArn=f'arn:{partition}:iam::{account_id}:role/{const.SOLUTION_NAME}RoleForAdmin-{region}',
        RoleSessionName="AssumeRoleSession1"
    )
    credentials = assumed_role_object['Credentials']

    glue = boto3.client(service_name='glue',
                        aws_access_key_id=credentials['AccessKeyId'],
                        aws_secret_access_key=credentials['SecretAccessKey'],
                        aws_session_token=credentials['SessionToken'],
                        region_name=region,
                        )
    glue_database_name = f'{const.SOLUTION_NAME}-{run_database.database_type}-{run_database.database_name}'
    if not is_structured:
        glue_database_name = f'{const.SOLUTION_NAME}-{DatabaseType.S3_UNSTRUCTURED.value}-{run_database.database_name}'
    elif run_database.database_type == DatabaseType.GLUE.value:
        glue_database_name = run_database.database_name
    next_token = None
    count = 0
    logger.info(base_time)
    while True:
        try:
            if next_token:
                response = glue.get_tables(DatabaseName=glue_database_name, NextToken=next_token)
            else:
                response = glue.get_tables(DatabaseName=glue_database_name)
        except Exception as e:
            logger.exception(e)
            return -1
        for table in response['TableList']:
            if table.get('Parameters', {}).get('classification', '') != 'UNKNOWN' and table['UpdateTime'] > base_time:
                count += 1
        next_token = response.get('NextToken')
        if not next_token:
            break
    return count


def __send_complete_run_database_message(job_id, run_id, run_database_id, account_id, region,
                                         database_type, database_name, over_write, state, message=""):
    event = {'Result': {'State': state, 'Message': message},
             'JobId': job_id,
             'RunId': run_id,
             'RunDatabaseId': run_database_id,
             'AccountId': account_id,
             'Region': region,
             'DatabaseType': database_type,
             'DatabaseName': database_name,
             'OverWrite': '1' if over_write else '0',
             }
    queue = sqs_resource.get_queue_by_name(QueueName=const.JOB_QUEUE_NAME)
    queue.send_message(MessageBody=json.dumps(event))


def __publish_job_completed(run_id: int):
    topic_arn = f'arn:{partition}:sns:{admin_region}:{admin_account_id}:{const.SOLUTION_NAME}-JobCompleted'
    sns = boto3.client('sns')

    job = crud.get_job_by_run_id(run_id)
    if not job:
        return

    message = {
        'JobId': job.id,
        'RunId': run_id,
        'Message': f'{job.name} has been completed.'
    }

    message_body = json.dumps(message)
    try:
        response = sns.publish(
            TopicArn=topic_arn,
            Message=message_body,
            Subject=message['Message'],
        )
    except Exception as e:
        logger.exception(e)


def complete_run(run_id: int):
    run_databases = crud.count_run_databases(run_id)
    # If there are running tasks, the state is running.
    for state, count in run_databases:
        if state in [RunDatabaseState.RUNNING.value, RunDatabaseState.PENDING.value]:
            logger.info("There are also running tasks.")
            return
    crud.complete_run(run_id)
    __publish_job_completed(run_id)


def complete_run_database(input_event):
    state = input_event["Result"]["State"]
    message = ""
    if "Message" in input_event["Result"]:
        message = input_event["Result"]["Message"]
    if state == RunDatabaseState.SUCCEEDED.value or state == RunDatabaseState.STOPPED.value:
        try:
            sync_job_detection_result(input_event["AccountId"],
                                      input_event["Region"],
                                      input_event["DatabaseType"],
                                      input_event["DatabaseName"],
                                      input_event["RunId"],
                                      input_event["OverWrite"] == "1",
                                      )
        except Exception:
            state = RunDatabaseState.FAILED.value
            message = traceback.format_exc()
            logger.exception("sync job detection result exception:%s" % message)
    run_database = crud.complete_run_database(input_event["RunDatabaseId"], state, message)
    job = crud.get_job(input_event["JobId"])
    if run_database and state == RunDatabaseState.SUCCEEDED.value and job.range == 1:
        crud.update_job_database_base_time(input_event["JobId"],
                                           input_event["AccountId"],
                                           input_event["Region"],
                                           input_event["DatabaseType"],
                                           input_event["DatabaseName"],
                                           run_database.start_time
                                            )
    logger.info(f'complete_run_database,JobId:{input_event["JobId"]},RunId:{input_event["RunId"]},DatabaseName:{input_event["DatabaseName"]}')


def check_pending_run_databases():
    run_databases = crud.get_pending_run_databases()
    if run_databases:
        __start_run_databases(run_databases)


def check_running_run_databases():
    run_databases = crud.get_running_run_databases()
    for run_database in run_databases:
        run_database_state, stop_time = __get_run_database_state_from_agent(run_database)
        logger.info(f"check running run,run id:{run_database.run_id},run database id:{run_database.id}"
                    f",account id:{run_database.account_id},region:{run_database.region}"
                    f",database type:{run_database.database_type},database name:{run_database.database_name}"
                    f",state:{run_database_state}")
        if run_database_state == RunDatabaseState.RUNNING.value.upper():
            continue
        state = run_database_state
        message = ""
        if run_database_state == RunDatabaseState.NOT_EXIST.value:
            message = 'Execution Does Not Exist'
        elif run_database_state == RunDatabaseState.SUCCEEDED.value.upper():
            if (datetime.datetime.now(pytz.timezone('UTC')) - stop_time).seconds < const.LAMBDA_MAX_RUNTIME:
                logger.info(f"run id:{run_database.run_id},run database id:{run_database.id} continue")
                continue
            state, message = __get_run_log(run_database, False)
        elif run_database_state == RunDatabaseState.FAILED.value.upper():
            state = RunDatabaseState.FAILED.value
            _, message = __get_run_log(run_database)
        elif run_database_state == RunDatabaseState.ABORTED.value.upper():
            state = RunDatabaseState.STOPPED.value

        job = crud.get_job_by_run_id(run_database.run_id)
        __send_complete_run_database_message(job.id, run_database.run_id, run_database.id, run_database.account_id,
                                             run_database.region, run_database.database_type,
                                             run_database.database_name,
                                             job.overwrite == 1, state, message)


def __get_run_database_state_from_agent(run_database: models.DiscoveryJobRunDatabase) -> (str, datetime.datetime):
    account_id = run_database.account_id
    region = run_database.region
    if need_change_account_id(run_database.database_type):
        account_id = admin_account_id
        region = admin_region
    client_sts = boto3.client('sts')
    assumed_role_object = client_sts.assume_role(
        RoleArn=f'arn:{partition}:iam::{account_id}:role/{const.SOLUTION_NAME}RoleForAdmin-{region}',
        RoleSessionName="AssumeRoleSession1"
    )
    credentials = assumed_role_object['Credentials']

    client_sfn = boto3.client('stepfunctions',
                              aws_access_key_id=credentials['AccessKeyId'],
                              aws_secret_access_key=credentials['SecretAccessKey'],
                              aws_session_token=credentials['SessionToken'],
                              region_name=region,
                              )

    try:
        response = client_sfn.describe_execution(
            executionArn=f'arn:{partition}:states:{region}:{account_id}:execution:{const.SOLUTION_NAME}-DiscoveryJob:{const.SOLUTION_NAME}-{run_database.run_id}-{run_database.id}-{run_database.uuid}',
        )
        return response.get("status"), response.get("stopDate")
    except client_sfn.exceptions.ExecutionDoesNotExist as e:
        return RunDatabaseState.NOT_EXIST.value, None


def __get_run_log(run_database: models.DiscoveryJobRunDatabase, error_log=True) -> (str, str):
    account_id = run_database.account_id
    region = run_database.region
    if need_change_account_id(run_database.database_type):
        account_id = admin_account_id
        region = admin_region
    client_sts = boto3.client('sts')
    assumed_role_object = client_sts.assume_role(
        RoleArn=f'arn:{partition}:iam::{account_id}:role/{const.SOLUTION_NAME}RoleForAdmin-{region}',
        RoleSessionName="AssumeRoleSession1"
    )
    credentials = assumed_role_object['Credentials']

    client_sfn = boto3.client('stepfunctions',
                              aws_access_key_id=credentials['AccessKeyId'],
                              aws_secret_access_key=credentials['SecretAccessKey'],
                              aws_session_token=credentials['SessionToken'],
                              region_name=region,
                              )
    max_results = 1 if error_log else 5
    try:
        response = client_sfn.get_execution_history(
            executionArn=f'arn:{partition}:states:{region}:{account_id}:execution:{const.SOLUTION_NAME}-DiscoveryJob:{const.SOLUTION_NAME}-{run_database.run_id}-{run_database.id}-{run_database.uuid}',
            reverseOrder=True,
            maxResults=max_results,
        )
    except client_sfn.exceptions.ExecutionDoesNotExist as e:
        return RunDatabaseState.NOT_EXIST.value, ""
    if error_log:
        if response["events"][0]["type"] == "ExecutionFailed":
            return RunDatabaseState.FAILED.value, response["events"][0]["executionFailedEventDetails"]["cause"]
        return RunDatabaseState.FAILED.value, ""
    else:
        parameters = json.loads(response["events"][4]["taskScheduledEventDetails"]["parameters"])
        result = parameters.get("MessageBody", {}).get("Result")
        if result:
            return result.get("State"), result.get("Message", "")
        return RunDatabaseState.SUCCEEDED.value, ""


def __get_cell_value(cell: dict):
    return cell.get("VarCharValue", "")


def generate_report(job_id: int, run_id: int, s3_client=None, key_name=None):
    logger.info(f"Gen job report,run id:{run_id}")
    if not s3_client:
        s3_client = boto3.client('s3')
    if not key_name:
        key_name = report_key_template % (job_id, run_id)
    job = crud.get_job(job_id)
    datasource_info = {}
    # Starting from version v1.1, a job only has one database_type
    if job.database_type.startswith(DatabaseType.JDBC.value):
        data_sources = list_resources_by_database_type(job.database_type).all()
        for data_source in data_sources:
            datasource_key = f"{job.database_type}-{data_source.instance_id}"
            datasource_info[datasource_key] = data_source

    run_result = __query_athena(sql_result % run_id)

    wb = Workbook()

    ws_s3_structured = wb.active
    ws_s3_structured.title = "Amazon S3(Structured)"
    ws_s3_unstructured = wb.create_sheet("Amazon S3(Unstructured)")
    ws_rds = wb.create_sheet("Amazon RDS")
    ws_jdbc = wb.create_sheet("JDBC")
    ws_glue = wb.create_sheet("Glue")
    ws_s3_structured.append(["account_id", "region", "bucket_name", "location", "column_name", "identifiers", "sample_data"])
    ws_s3_unstructured.append(["account_id", "region", "bucket_name", "location", "identifiers", "sample_data"])
    ws_rds.append(["account_id", "region", "instance_name", "table_name", "column_name", "identifiers", "sample_data"])
    ws_jdbc.append(["type", "account_id", "region", "instance_name", "description", "jdbc_url", "table_name", "column_name", "identifiers", "sample_data"])
    ws_glue.append(["account_id", "region", "database_name", "table_name", "column_name", "identifiers", "sample_data"])

    for row in run_result[1:]:
        row_result = [__get_cell_value(cell) for cell in row]
        database_type = row_result[0]
        del row_result[0]
        if database_type == DatabaseType.S3.value:
            ws_s3_structured.append(row_result)
        elif database_type == DatabaseType.S3_UNSTRUCTURED.value:
            del row_result[4]  # Delete column_name field
            ws_s3_unstructured.append(row_result)
        elif database_type == DatabaseType.GLUE.value:
            ws_glue.append(row_result)
        elif database_type.startswith(DatabaseType.JDBC.value):
            datasource_key = f"{database_type}-{row_result[2]}"
            data_source = datasource_info[datasource_key]
            row_result[3:3] = [data_source.description, data_source.jdbc_connection_url]
            row_result.insert(0, database_type[5:])
            ws_jdbc.append(row_result)
        else:  # RDS
            ws_rds.append(row_result)

    error_result = __query_athena(sql_error % run_id)
    if len(error_result) > 1:
        ws_failed = wb.create_sheet("Detect failed tables")
        ws_failed.append(["account_id", "region", "database_type", "database_name", "table_name", "error_message"])
        for row in error_result[1:]:
            row_result = [__get_cell_value(cell) for cell in row]
            ws_failed.append(row_result)

    filename = NamedTemporaryFile().name
    wb.save(filename)
    s3_client.upload_file(filename, admin_bucket_name, key_name)
    os.remove(filename)


def __check_file_existence(s3_client, key_name):
    try:
        s3_client.head_object(Bucket=admin_bucket_name, Key=key_name)
        return True
    except Exception as e:
        if e.response['Error']['Code'] == '404':
            return False
        else:
            logger.info(e)
            return False


def get_report_url(job_id: int, run_id: int):
    s3_client = boto3.client('s3')
    key_name = report_key_template % (job_id, run_id)
    if not __check_file_existence(s3_client, key_name):
        generate_report(job_id, run_id, s3_client, key_name)
    method_parameters = {'Bucket': admin_bucket_name, 'Key': key_name}
    pre_url = s3_client.generate_presigned_url(
        ClientMethod="get_object",
        Params=method_parameters,
        ExpiresIn=60
    )
    return pre_url


def __query_athena(sql: str):
    logger.debug(sql)
    client = boto3.client("athena")
    queryStart = client.start_query_execution(
        QueryString=sql,
        QueryExecutionContext={
            "Database": const.JOB_RESULT_DATABASE_NAME,
            "Catalog": "AwsDataCatalog",
        },
        ResultConfiguration={"OutputLocation": f"s3://{admin_bucket_name}/athena-output/"},
    )

    query_execution_id = queryStart["QueryExecutionId"]
    while True:
        response = client.get_query_execution(QueryExecutionId=query_execution_id)
        query_execution_status = response["QueryExecution"]["Status"]["State"]
        if query_execution_status == AthenaQueryState.SUCCEEDED.value:
            break
        if query_execution_status == AthenaQueryState.FAILED.value:
            logger.exception(response)
            raise Exception("Query Asset STATUS:" + response["QueryExecution"]["Status"]["StateChangeReason"])
        else:
            time.sleep(1)

    results_paginator = client.get_paginator('get_query_results')
    results_iterator = results_paginator.paginate(
        QueryExecutionId=query_execution_id,
        PaginationConfig={
            'PageSize': 1000
        }
    )

    first_page = True
    result = []
    for results_page in results_iterator:
        tmp_result = [x["Data"] for x in results_page["ResultSet"]["Rows"]]
        if first_page:
            first_page = False
            result += tmp_result
        else:
            result += tmp_result[1:]
    return result


def get_template_snapshot_url(run_id: int):
    job_run = crud.get_run(run_id)
    if job_run.template_id is None or job_run.template_snapshot_no is None:
        raise BizException(MessageEnum.DISCOVERY_RUN_NON_EXIST_TEMPLATE_SNAPSHOT.get_code(),
                           MessageEnum.DISCOVERY_RUN_NON_EXIST_TEMPLATE_SNAPSHOT.get_msg())
    s3_client = boto3.client('s3')
    key_name = f"template/template-{job_run.template_id}-{job_run.template_snapshot_no}.json"
    method_parameters = {'Bucket': admin_bucket_name, 'Key': key_name}
    pre_url = s3_client.generate_presigned_url(
        ClientMethod="get_object",
        Params=method_parameters,
        ExpiresIn=60
    )
    return pre_url


def can_delete_account(account_id: str, region: str):
    db_count = crud.count_account_run_job(account_id, region)
    return db_count == 0


def delete_account(account_id: str, region: str):
    if can_delete_account(account_id, region):
        crud.delete_account(account_id, region)
    else:
        raise BizException(MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_ACCOUNT.get_code(),
                           MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_ACCOUNT.get_msg())


def can_delete_database(account_id: str, region: str, database_type: str, database_name: str):
    db_count = crud.count_database_run_job(account_id, region, database_type, database_name)
    return db_count == 0


def delete_database(account_id: str, region: str, database_type: str, database_name: str):
    if can_delete_database(account_id, region, database_type, database_name):
        crud.delete_database(account_id, region, database_type, database_name)
    else:
        raise BizException(MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_code(),
                           MessageEnum.DISCOVERY_JOB_CAN_NOT_DELETE_DATABASE.get_msg())
